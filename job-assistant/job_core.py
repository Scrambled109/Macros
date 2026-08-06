"""Testable workflow services for the Engineering Job Assistant."""

from __future__ import annotations

import csv
import filecmp
import getpass
import json
import os
import re
import shutil
import sys
import uuid
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Sequence

import openpyxl

MANIFEST_NAME = "job_manifest.json"
MANIFEST_VERSION = 4
SETTINGS_VERSION = 3
ASSISTANT_DIR = "_JOB_ASSISTANT"

STAGES = [
    ("bom", "Create Parts List from BOP/BOM"),
    ("dxf", "Review and prepare cut files"),
    ("plate_model", "Create automatic plate models"),
    ("autobom", "Apply modified Parts List properties"),
    ("comparison", "Compare production data"),
]

# Older manifests retain these completed/manual checkpoints for audit history,
# but they are no longer shown as work performed by the assistant.
REMOVED_STAGES = {
    "manual_model": "Create shape and specialized models",
    "assembly": "Build and review assembly",
    "nesting": "Create nesting outputs",
    "final": "Review discrepancies and finalize",
}

WORKSPACE_NAMES = {
    "assistant": ASSISTANT_DIR,
    "source_copies": f"{ASSISTANT_DIR}/Source Copies",
    "working": f"{ASSISTANT_DIR}/Working",
    "staging": f"{ASSISTANT_DIR}/Staging",
    "logs": f"{ASSISTANT_DIR}/Logs",
    "backups": f"{ASSISTANT_DIR}/Backups",
    "reports": f"{ASSISTANT_DIR}/Reports",
}

STAGE_GUIDANCE = {
    "bom": {
        "need": "The received BOP/BOM and the standard Parts List template.",
        "action": "Select the BOP. Confirm the recommended Parts List name and destination.",
        "changes": "The BOP is copied to Source Copies; the converter writes the selected output.",
        "tool": "Existing BOM converter.",
        "review": "Open the Parts List and verify part numbers, quantities, material, and dimensions.",
    },
    "dxf": {
        "need": "The folder containing received DXFs and shape sketches.",
        "action": "Review every proposed file. DXFs start selected; DWGs start excluded.",
        "changes": "Selected files are copied into numbered working folder 001. Originals are untouched.",
        "tool": "Python DXF orchestrator, run against working copies.",
        "review": "Review every generated DWG and the orchestrator logs; detected bevel drawings are identified by the (B) filename suffix.",
    },
    "plate_model": {
        "need": "Reviewed, prepared DWGs.",
        "action": "Select one material/thickness DWG folder and confirm its extrusion thickness.",
        "changes": "The assistant writes per-run settings, reuses the active SolidWorks instance (or starts one once), and runs each selected folder through the background controller.",
        "tool": "Compiled CAD batch converter macro through the reusable SolidWorks Python runner.",
        "review": "The result folder opens after every batch. Inspect geometry, thickness, markings, and BatchLog.txt before selecting the next folder.",
    },
    "autobom": {
        "need": "The reviewed assembly, writable part files, and the modified Parts List workbook.",
        "action": "Run the modified Parts List property macro only after making a recoverable model copy.",
        "changes": "The macro maps spreadsheet columns to Description/Raw_Material, updates properties, and saves part files.",
        "tool": "(MOD)2(SECONDARY) SolidWorks macro.",
        "review": "Review mapped properties, unmatched parts, save failures, and skipped files.",
    },
    "comparison": {
        "need": "Parts List CSV, SolidWorks CSV, and nesting export folder.",
        "action": "Select inputs and run the production comparison.",
        "changes": "Reports are written below assistant Reports; production inputs are read-only.",
        "tool": "Existing production comparison tool.",
        "review": "Open Excel/HTML reports and resolve errors; launch alone is not a pass.",
    },
}


def inferred_plate_thickness(source: Path) -> float | None:
    """Infer inches from the orchestrator's leading-mils folder convention."""
    match = re.match(r"^(\d+)-", source.name)
    return int(match.group(1)) / 1000 if match else None

class JobError(RuntimeError):
    """A problem that can be explained and corrected by the operator."""


@dataclass(frozen=True)
class Check:
    level: str
    code: str
    message: str
    correction: str = ""


@dataclass(frozen=True)
class DrawingCandidate:
    path: Path
    selected: bool
    kind: str


@dataclass(frozen=True)
class OutputMoveItem:
    source: Path
    destination: Path
    category: str
    run_name: str
    conflict: bool


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def operator_name() -> str:
    return os.environ.get("USERNAME") or getpass.getuser() or "unknown"


def safe_name(value: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "-", value.strip())
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" .")
    if not cleaned:
        raise JobError("Enter a usable job number, name, or revision.")
    return cleaned


def suggest_job_number(root: Path) -> str:
    """Suggest leading digits from the engineering folder or one of its parents."""
    for candidate in (root, *root.parents):
        match = re.match(r"\s*(\d+)", candidate.name)
        if match:
            return match.group(1)
    return ""


def settings_path() -> Path:
    if sys.platform == "win32":
        base = Path(os.environ.get("LOCALAPPDATA", Path.home() / "AppData/Local"))
    else:
        base = Path(os.environ.get("XDG_CONFIG_HOME", Path.home() / ".config"))
    return base / "EngineeringJobAssistant" / "settings.json"


def default_settings(repo_root: Path | None = None) -> dict[str, Any]:
    return {
        "settings_version": SETTINGS_VERSION,
        "macros_repo": str(repo_root.resolve()) if repo_root else "",
        "parts_list_template": "",
        "default_jobs_parent": "",
        "autocad_console": "",
        "autocad_workers": 2,
        "solidworks_executable": "",
    }


def load_settings(
    path: Path | None = None, repo_root: Path | None = None
) -> dict[str, Any]:
    target = path or settings_path()
    settings = default_settings(repo_root)
    if not target.exists():
        return settings
    try:
        loaded = json.loads(target.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise JobError(f"Could not read local settings at {target}: {exc}") from exc
    if not isinstance(loaded, dict):
        raise JobError(f"Local settings at {target} are not a JSON object.")
    settings.update({key: value for key, value in loaded.items() if key in settings})
    try:
        settings["autocad_workers"] = max(
            1, min(4, int(settings.get("autocad_workers", 2)))
        )
    except (TypeError, ValueError):
        settings["autocad_workers"] = 2
    settings["settings_version"] = SETTINGS_VERSION
    return settings


def _atomic_json(data: dict[str, Any], target: Path) -> Path:
    target.parent.mkdir(parents=True, exist_ok=True)
    temporary = target.with_name(f".{target.name}.{uuid.uuid4().hex}.tmp")
    try:
        with temporary.open("w", encoding="utf-8", newline="\n") as handle:
            json.dump(data, handle, indent=2)
            handle.write("\n")
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, target)
    finally:
        if temporary.exists():
            temporary.unlink()
    return target


def save_settings(settings: dict[str, Any], path: Path | None = None) -> Path:
    clean = default_settings()
    clean.update({key: value for key, value in settings.items() if key in clean})
    try:
        clean["autocad_workers"] = max(
            1, min(4, int(clean.get("autocad_workers", 2)))
        )
    except (TypeError, ValueError):
        clean["autocad_workers"] = 2
    clean["settings_version"] = SETTINGS_VERSION
    return _atomic_json(clean, path or settings_path())


def _stage(label: str) -> dict[str, Any]:
    return {
        "label": label,
        "status": "not_started",
        "started_at": None,
        "completed_at": None,
        "reviewed": False,
        "notes": "",
        "warnings_acknowledged": [],
        "artifacts": [],
    }


def new_manifest(
    job_number: str,
    job_name: str,
    revision: str,
    engineering_root: Path,
    model_folder: Path,
    cut_folder: Path,
) -> dict[str, Any]:
    root = engineering_root.resolve()
    manifest = {
        "manifest_version": MANIFEST_VERSION,
        "job": {
            "number": safe_name(job_number),
            "name": safe_name(job_name or job_number),
            "revision": safe_name(revision),
            "created_at": utc_now(),
            "updated_at": utc_now(),
        },
        "root": str(root),
        "paths": {
            "engineering_root": str(root),
            "model_3d": str(model_folder.resolve()),
            "cut_files": str(cut_folder.resolve()),
            "part_checking": "",
            "nesting": "",
            "forming": "",
        },
        "workspace": {
            key: str((root / value).resolve()) for key, value in WORKSPACE_NAMES.items()
        },
        "stages": {key: _stage(label) for key, label in STAGES},
        "legacy_stages": {},
        "output_moves": [],
        "events": [],
        "recent_files": [],
        "comparison": None,
    }
    record_event(manifest, "job_attached", root=str(root))
    return manifest


def setup_job(
    engineering_root: Path,
    model_folder: Path,
    cut_folder: Path,
    job_number: str,
    job_name: str = "",
    revision: str = "A",
) -> Path:
    for label, folder in (
        ("Engineering Process", engineering_root),
        ("3D Model", model_folder),
        ("Cut Files", cut_folder),
    ):
        if not folder.is_dir():
            raise JobError(f"The selected {label} folder does not exist: {folder}")
    manifest_path = engineering_root / ASSISTANT_DIR / MANIFEST_NAME
    if manifest_path.exists():
        raise JobError(
            f"This job already has an assistant manifest: {manifest_path}. "
            "Open it instead of replacing its audit history."
        )
    manifest = new_manifest(
        job_number, job_name, revision, engineering_root, model_folder, cut_folder
    )
    ensure_workspace(manifest)
    return save_manifest(manifest)


def ensure_workspace(manifest: dict[str, Any]) -> None:
    for path in manifest["workspace"].values():
        Path(path).mkdir(parents=True, exist_ok=True)


def migrate_manifest(raw: dict[str, Any], source: Path | None = None) -> dict[str, Any]:
    version = raw.get("manifest_version", 1)
    if not isinstance(version, int) or version > MANIFEST_VERSION:
        raise JobError(
            f"This manifest uses version {version}; this assistant supports up to version {MANIFEST_VERSION}. Update the assistant before opening it."
        )
    if not raw.get("root") or not isinstance(raw.get("job"), dict):
        raise JobError(
            "The job manifest is incomplete: root and job details are required."
        )
    job = raw["job"]
    if not str(job.get("number", "")).strip():
        raise JobError("The job manifest is incomplete: job number is missing.")
    if not str(job.get("revision", "")).strip():
        raise JobError("The job manifest is incomplete: revision is missing.")
    job.setdefault("name", job["number"])
    job.setdefault("created_at", utc_now())
    job.setdefault("updated_at", utc_now())
    root = Path(raw["root"])
    old_paths = raw.setdefault("paths", {})
    model = old_paths.get("model_3d") or old_paths.get("cad_models") or root
    cut = old_paths.get("cut_files") or root
    paths = {
        "engineering_root": str(root),
        "model_3d": str(model),
        "cut_files": str(cut),
        "part_checking": old_paths.get("part_checking", ""),
        "nesting": old_paths.get("nesting") or old_paths.get("nests", ""),
        "forming": old_paths.get("forming", ""),
    }
    workspace = {
        key: str((root / value).resolve()) for key, value in WORKSPACE_NAMES.items()
    }
    for key, old_key in (("logs", "logs"), ("reports", "reports")):
        if old_paths.get(old_key):
            workspace[key] = old_paths[old_key]
    raw["paths"], raw["workspace"] = paths, workspace
    stages = raw.setdefault("stages", {})
    legacy_stages = raw.setdefault("legacy_stages", {})
    for key, label in REMOVED_STAGES.items():
        if key not in stages:
            continue
        legacy_item = dict(stages.pop(key))
        legacy_item.setdefault("label", label)
        legacy_stages.setdefault(key, legacy_item)
    for key, label in STAGES:
        base = _stage(label)
        base.update(stages.get(key, {}))
        base["label"] = label
        stages[key] = base
    raw.setdefault("events", [])
    raw.setdefault("recent_files", [])
    raw.setdefault("output_moves", [])
    raw.setdefault("comparison", None)
    if version != MANIFEST_VERSION:
        record_event(
            raw,
            "manifest_migrated",
            from_version=version,
            to_version=MANIFEST_VERSION,
            source=str(source or ""),
        )
    raw["manifest_version"] = MANIFEST_VERSION
    return raw


def load_manifest(path: Path) -> dict[str, Any]:
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise JobError(f"Could not read manifest {path}: {exc}") from exc
    if not isinstance(raw, dict):
        raise JobError("The selected manifest is not a JSON object.")
    return migrate_manifest(raw, path)


def save_manifest(manifest: dict[str, Any], path: Path | None = None) -> Path:
    manifest["manifest_version"] = MANIFEST_VERSION
    manifest["job"]["updated_at"] = utc_now()
    return _atomic_json(
        manifest, path or Path(manifest["root"]) / ASSISTANT_DIR / MANIFEST_NAME
    )


def record_event(
    manifest: dict[str, Any], event_type: str, **details: Any
) -> dict[str, Any]:
    event = {"at": utc_now(), "user": operator_name(), "type": event_type, **details}
    manifest.setdefault("events", []).append(event)
    return event


def set_optional_path(manifest: dict[str, Any], key: str, path: Path) -> None:
    if key not in {"part_checking", "nesting", "forming"}:
        raise JobError(f"Unknown optional job folder: {key}")
    if not path.is_dir():
        raise JobError(f"The selected folder does not exist: {path}")
    manifest["paths"][key] = str(path.resolve())
    record_event(manifest, "job_path_changed", path_key=key, path=str(path.resolve()))


def copy_source(source: Path, destination_folder: Path) -> Path:
    if not source.is_file():
        raise JobError(f"The selected source file does not exist: {source}")
    destination_folder.mkdir(parents=True, exist_ok=True)
    target = destination_folder / source.name
    if target.exists():
        stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        target = (
            destination_folder
            / f"{source.stem}_{stamp}_{uuid.uuid4().hex[:6]}{source.suffix}"
        )
    shutil.copy2(source, target)
    return target


def discover_drawings(folder: Path) -> list[DrawingCandidate]:
    if not folder.is_dir():
        raise JobError(f"The incoming drawing folder does not exist: {folder}")
    result = []
    for path in sorted(
        (item for item in folder.iterdir() if item.is_file()),
        key=lambda item: item.name.lower(),
    ):
        suffix = path.suffix.lower()
        if suffix == ".dxf":
            result.append(DrawingCandidate(path.resolve(), True, "DXF plate candidate"))
        elif suffix == ".dwg":
            result.append(
                DrawingCandidate(
                    path.resolve(), False, "DWG likely manual shape sketch"
                )
            )
    return result


def prepare_dxf_workspace(
    manifest: dict[str, Any], selected: Sequence[Path], parts_list_csv: Path
) -> Path:
    if not selected:
        raise JobError("Select at least one drawing before preparing the workspace.")
    base = Path(manifest["workspace"]["working"]) / "DXF Orchestrator"
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    run = base / stamp
    sequence = 2
    while run.exists():
        run = base / f"{stamp}-{sequence}"
        sequence += 1
    numbered = run / "001"
    numbered.mkdir(parents=True, exist_ok=False)
    if not parts_list_csv.is_file():
        raise JobError(f"The DXF Parts List CSV does not exist: {parts_list_csv}")
    shutil.copy2(parts_list_csv, run / "Parts List.csv")
    seen: set[str] = set()
    for source in selected:
        if not source.is_file():
            raise JobError(f"A selected drawing no longer exists: {source}")
        key = source.name.casefold()
        if key in seen:
            raise JobError(f"Two selected files have the same name: {source.name}")
        seen.add(key)
        shutil.copy2(source, numbered / source.name)
    record_event(manifest, "dxf_workspace_prepared", run=str(run), files=len(selected))
    manifest["stages"]["dxf"]["status"] = "ready"
    manifest["stages"]["dxf"]["workspace"] = str(run)
    return run


def export_parts_list_csv(workbook_path: Path, destination: Path) -> Path:
    """Export the orchestrator columns from a generated Parts List workbook."""
    if not workbook_path.is_file():
        raise JobError(f"The generated Parts List does not exist: {workbook_path}")
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        header_row = None
        columns: dict[str, int] = {}
        aliases = {
            "PART NUMBER": "PartNumber",
            "QUANTITY": "Quantity",
            "TOTAL QUANTITY": "Quantity",
            "THICKNESS SHAPE": "Thickness",
            "THICKNESS/SHAPE": "Thickness",
            "MATERIAL TYPE": "Material",
            "MATERIAL": "Material",
            "DESCRIPTION": "Description",
        }
        for sheet in workbook.worksheets:
            for row_number, row in enumerate(
                sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 50)), start=1
            ):
                found: dict[str, int] = {}
                for index, cell in enumerate(row):
                    normalized = re.sub(
                        r"[^A-Z0-9]+", " ", str(cell.value or "").upper()
                    ).strip()
                    canonical = aliases.get(normalized)
                    if canonical and canonical not in found:
                        found[canonical] = index
                required = {"PartNumber", "Thickness", "Material", "Description"}
                if required.issubset(found):
                    header_row = (sheet, row_number)
                    columns = found
                    break
            if header_row:
                break
        if not header_row:
            raise JobError(
                "Could not find PART NUMBER, DESCRIPTION, THICKNESS/SHAPE, and "
                "MATERIAL TYPE columns in the generated Parts List."
            )

        sheet, row_number = header_row
        destination.parent.mkdir(parents=True, exist_ok=True)
        with destination.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(
                handle,
                fieldnames=["PartNumber", "Quantity", "Thickness", "Material"],
            )
            writer.writeheader()
            for row in sheet.iter_rows(min_row=row_number + 1, values_only=True):
                description = re.sub(
                    r"[^A-Z0-9]+",
                    " ",
                    str(row[columns["Description"]] or "").upper(),
                ).strip()
                if description not in {"PL", "PLATE"}:
                    continue
                part = row[columns["PartNumber"]]
                if part is None or not str(part).strip():
                    continue
                writer.writerow(
                    {
                        "PartNumber": str(part).strip(),
                        "Quantity": str(
                            row[columns["Quantity"]]
                            if "Quantity" in columns and row[columns["Quantity"]] is not None
                            else "1"
                        ).strip(),
                        "Thickness": str(row[columns["Thickness"]] or "").strip(),
                        "Material": str(row[columns["Material"]] or "").strip(),
                    }
                )
    finally:
        workbook.close()
    return destination


def command_bom(
    python: Path | str,
    repo: Path,
    source_copy: Path,
    output: Path,
    template: Path,
    tool_executable: Path | None = None,
) -> list[str]:
    prefix = (
        [str(tool_executable)]
        if tool_executable
        else [str(python), str(repo / "data-tools/bom-converter/bom_converter.py")]
    )
    return [*prefix, "--gui", str(source_copy), str(output), str(template)]


def existing_working_directory(path: Path | str | None) -> Path | None:
    """Return *path* only when it is safe to pass to ``subprocess`` as cwd."""
    if not path:
        return None
    candidate = Path(path)
    return candidate if candidate.is_dir() else None


def command_dxf(
    python: Path | str,
    repo: Path,
    workspace: Path,
    parts_list_csv: Path,
    autocad_console: Path | str | None = None,
    tool_executable: Path | None = None,
    workers: int = 2,
) -> list[str]:
    command = (
        [str(tool_executable)]
        if tool_executable
        else [
            str(python),
            "-u",
            str(repo / "autocad/dxf-orchestrator/Master_Orchestrator.py"),
        ]
    )
    command.extend(["--workspace", str(workspace), "--parts-list", str(parts_list_csv)])
    command.extend(["--workers", str(max(1, min(4, int(workers))))])
    if autocad_console:
        command.extend(["--acad-console-path", str(autocad_console)])
    return command


def command_comparison(
    python: Path | str,
    repo: Path,
    nests: Path,
    parts: Path,
    solidworks: Path,
    output: Path,
    tool_executable: Path | None = None,
) -> list[str]:
    prefix = (
        [str(tool_executable)]
        if tool_executable
        else [
            str(python),
            "-u",
            str(repo / "data-tools/production-comparison/compare_production_parts.py"),
        ]
    )
    return [
        *prefix,
        "--nests",
        str(nests),
        "--parts",
        str(parts),
        "--solidworks",
        str(solidworks),
        "--output",
        str(output),
        "--no-open",
    ]


def acknowledge_override(
    manifest: dict[str, Any],
    stage: str,
    checks: Iterable[Check],
    action: str = "continue_anyway",
) -> None:
    acknowledged = [asdict(check) for check in checks]
    manifest["stages"][stage]["warnings_acknowledged"].extend(acknowledged)
    record_event(
        manifest,
        "warning_overridden",
        stage=stage,
        action=action,
        checks=[item["code"] for item in acknowledged],
    )


def start_stage(
    manifest: dict[str, Any], stage: str, warnings: Iterable[Check] = ()
) -> None:
    item = manifest["stages"][stage]
    if item["status"] == "complete":
        raise JobError("Reopen this completed stage before running it again.")
    item["status"], item["started_at"] = (
        "in_progress",
        item.get("started_at") or utc_now(),
    )
    if warnings:
        acknowledge_override(manifest, stage, warnings)
    record_event(manifest, "stage_started", stage=stage)


def mark_needs_review(
    manifest: dict[str, Any], stage: str, message: str, artifacts: Iterable[Path] = ()
) -> None:
    item = manifest["stages"][stage]
    item["status"], item["notes"] = "needs_review", message
    for artifact in artifacts:
        if artifact.is_file():
            record_artifact(manifest, stage, artifact)
    record_event(manifest, "stage_needs_review", stage=stage, message=message)


def complete_stage(manifest: dict[str, Any], stage: str, notes: str) -> None:
    if manifest["stages"][stage]["status"] not in {
        "in_progress",
        "needs_review",
        "warning",
    }:
        raise JobError(
            "Start the stage and review its results before marking it complete."
        )
    if not notes.strip():
        raise JobError("Enter a short review note before completing the stage.")
    item = manifest["stages"][stage]
    item.update(
        status="complete", reviewed=True, completed_at=utc_now(), notes=notes.strip()
    )
    record_event(manifest, "stage_completed", stage=stage, notes=notes.strip())


def reopen_stage(
    manifest: dict[str, Any], stage: str, reason: str = "Operator reopened stage"
) -> None:
    item = manifest["stages"][stage]
    item.update(status="in_progress", reviewed=False, completed_at=None, notes=reason)
    record_event(manifest, "stage_reopened", stage=stage, reason=reason)


def change_revision(manifest: dict[str, Any], revision: str) -> None:
    old, new = manifest["job"]["revision"], safe_name(revision)
    if old != new:
        manifest["job"]["revision"] = new
        record_event(manifest, "revision_changed", old=old, new=new)


def record_artifact(manifest: dict[str, Any], stage: str, path: Path) -> dict[str, Any]:
    if not path.is_file():
        raise JobError(f"Artifact does not exist: {path}")
    artifact = {
        "path": str(path.resolve()),
        "name": path.name,
        "size": path.stat().st_size,
        "recorded_at": utc_now(),
        "revision": manifest["job"]["revision"],
        "user": operator_name(),
    }
    items = manifest["stages"][stage]["artifacts"]
    items[:] = [item for item in items if item.get("path") != artifact["path"]]
    items.append(artifact)
    manifest["recent_files"] = (
        [artifact]
        + [
            item
            for item in manifest.get("recent_files", [])
            if item.get("path") != artifact["path"]
        ]
    )[:10]
    record_event(
        manifest,
        "artifact_recorded",
        stage=stage,
        path=artifact["path"],
        revision=artifact["revision"],
    )
    return artifact


def stage_checks(manifest: dict[str, Any], stage: str) -> list[Check]:
    checks: list[Check] = []
    for key in ("engineering_root", "model_3d", "cut_files"):
        if not Path(manifest["paths"][key]).is_dir():
            checks.append(
                Check(
                    "block",
                    f"missing_{key}",
                    f"Selected job folder is unavailable: {manifest['paths'][key]}",
                    "Reconnect the drive or update the selected path.",
                )
            )
    keys = [key for key, _ in STAGES]
    position = keys.index(stage)
    if position and manifest["stages"][keys[position - 1]]["status"] != "complete":
        checks.append(
            Check(
                "warning",
                "previous_stage_incomplete",
                f"Previous stage '{manifest['stages'][keys[position - 1]]['label']}' is not complete.",
                "Review the dependency, or Continue Anyway if the sequence is intentionally different.",
            )
        )
    optional = {"nesting": "nesting", "comparison": "nesting"}.get(stage)
    if optional and not manifest["paths"].get(optional):
        checks.append(
            Check(
                "warning",
                f"select_{optional}",
                f"Select the {optional.title()} folder before this stage.",
                "Use Set Optional Folder on the dashboard.",
            )
        )
    return checks or [Check("pass", "ready", "This stage is ready to begin.")]


def recommended_next_action(manifest: dict[str, Any]) -> str:
    for key, label in STAGES:
        status = manifest["stages"][key]["status"]
        if status == "needs_review":
            return f"Review the results for: {label}."
        if status != "complete":
            return f"Next: {label}. Select the stage to see what is needed."
    return "All assistant steps are complete. Move completed outputs when ready."


def dashboard_warnings(manifest: dict[str, Any]) -> list[str]:
    """Return concise, current warnings suitable for the dashboard."""
    warnings: list[str] = []
    for label, key in (
        ("Engineering Process", "engineering_root"),
        ("3D Model", "model_3d"),
        ("Cut Files", "cut_files"),
    ):
        value = manifest["paths"].get(key, "")
        if not value or not Path(value).is_dir():
            warnings.append(f"{label} folder is unavailable: {value or 'not selected'}")
    for key, _label in STAGES:
        item = manifest["stages"][key]
        if item["status"] == "warning":
            warnings.append(
                f"{item['label']}: {item.get('notes') or 'action required'}"
            )
    comparison = manifest.get("comparison") or {}
    if comparison.get("status") == "action_required":
        warnings.append(
            "Production comparison has "
            f"{comparison.get('errors', 0)} actionable discrepancy item(s)."
        )
    return warnings


_CUT_FILE_PART_SUFFIX = re.compile(
    r"^(?P<part>.+)_[^_]+_\\d+(?:\\(B\\))?$", re.IGNORECASE
)


def solidworks_part_stem(cut_file_stem: str) -> str:
    """Return the original part number from an orchestrator cut-file stem."""
    match = _CUT_FILE_PART_SUFFIX.fullmatch(cut_file_stem)
    return match.group("part") if match else cut_file_stem


def normalize_plate_model_filenames(output_folder: Path) -> list[Path]:
    """Remove orchestrator material/quantity suffixes from generated parts.

    Only .SLDPRT files below the SolidWorks output folder are changed. Existing
    destination names and duplicate normalized names stop the operation before
    anything is renamed, so a generated model is never silently overwritten.
    """
    if not output_folder.is_dir():
        raise JobError(f"SolidWorks output folder does not exist: {output_folder}")

    plan: list[tuple[Path, Path]] = []
    destinations: dict[str, Path] = {}
    for source in sorted(
        (
            path
            for path in output_folder.rglob("*")
            if path.is_file() and path.suffix.casefold() == ".sldprt"
        ),
        key=lambda path: str(path).casefold(),
    ):
        target = source.with_name(f"{solidworks_part_stem(source.stem)}{source.suffix}")
        if target == source:
            continue
        key = os.path.normcase(str(target.resolve()))
        previous = destinations.get(key)
        if previous is not None:
            raise JobError(
                "Two generated SolidWorks parts would have the same original "
                f"name: {previous.name} and {source.name} -> {target.name}"
            )
        if target.exists():
            raise JobError(
                f"Cannot rename {source.name} to {target.name} because that file "
                "already exists. Review the generated models before moving them."
            )
        destinations[key] = source
        plan.append((source, target))

    renamed: list[Path] = []
    try:
        for source, target in plan:
            source.rename(target)
            renamed.append(target)
    except OSError as exc:
        for target in reversed(renamed):
            source = next(old for old, new in plan if new == target)
            try:
                target.rename(source)
            except OSError:
                pass
        raise JobError(f"Could not normalize SolidWorks part filenames: {exc}") from exc
    return renamed


def plan_completed_outputs(manifest: dict[str, Any]) -> list[OutputMoveItem]:
    """Find reviewed cut files and plate models that are ready for production.

    Cut-file material folders are merged below the selected Cut Files folder.
    Plate models are moved directly into the selected 3D Model folder so an
    assembly does not need to search through thickness-group subfolders.
    """

    planned: list[OutputMoveItem] = []
    seen_destinations: set[str] = set()

    def add(source: Path, destination: Path, category: str, run_name: str) -> None:
        key = os.path.normcase(str(destination.resolve()))
        conflict = destination.exists() or key in seen_destinations
        seen_destinations.add(key)
        planned.append(
            OutputMoveItem(
                source.resolve(),
                destination.resolve(),
                category,
                run_name,
                conflict,
            )
        )

    if manifest["stages"]["dxf"]["status"] == "complete":
        orchestrator_root = (
            Path(manifest["workspace"]["working"]) / "DXF Orchestrator"
        )
        cut_root = Path(manifest["paths"]["cut_files"])
        if orchestrator_root.is_dir():
            for run in sorted(
                (path for path in orchestrator_root.iterdir() if path.is_dir()),
                key=lambda path: path.name.casefold(),
            ):
                for group in sorted(
                    (
                        path
                        for path in run.iterdir()
                        if path.is_dir() and re.match(r"^\d+-\S", path.name)
                    ),
                    key=lambda path: path.name.casefold(),
                ):
                    for source in sorted(
                        (
                            path
                            for path in group.iterdir()
                            if path.is_file() and path.suffix.casefold() == ".dwg"
                        ),
                        key=lambda path: path.name.casefold(),
                    ):
                        add(
                            source,
                            cut_root / group.name / source.name,
                            "cut_file",
                            run.name,
                        )

    if manifest["stages"]["plate_model"]["status"] == "complete":
        plate_root = (
            Path(manifest["workspace"]["staging"]) / "SolidWorks Parts"
        )
        model_root = Path(manifest["paths"]["model_3d"])
        if plate_root.is_dir():
            for source in sorted(
                (
                    path
                    for path in plate_root.rglob("*")
                    if path.is_file() and path.suffix.casefold() == ".sldprt"
                ),
                key=lambda path: str(path).casefold(),
            ):
                relative = source.relative_to(plate_root)
                run_name = relative.parts[0] if len(relative.parts) > 1 else "plates"
                add(
                    source,
                    model_root / source.name,
                    "plate_model",
                    run_name,
                )

    return planned


def _remove_empty_output_parents(path: Path, stop: Path) -> None:
    current = path.parent
    stop = stop.resolve()
    while current != stop and current.is_relative_to(stop):
        try:
            current.rmdir()
        except OSError:
            break
        current = current.parent


def move_completed_outputs(
    manifest: dict[str, Any], items: Iterable[OutputMoveItem]
) -> list[dict[str, Any]]:
    """Safely move completed output files with automatic conflict backups."""

    orchestrator_root = (
        Path(manifest["workspace"]["working"]) / "DXF Orchestrator"
    ).resolve()
    plate_root = (
        Path(manifest["workspace"]["staging"]) / "SolidWorks Parts"
    ).resolve()
    cut_root = Path(manifest["paths"]["cut_files"]).resolve()
    model_root = Path(manifest["paths"]["model_3d"]).resolve()
    roots = {
        "cut_file": (orchestrator_root, cut_root),
        "plate_model": (plate_root, model_root),
    }
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_root = Path(manifest["workspace"]["backups"]) / stamp
    results: list[dict[str, Any]] = []

    for item in items:
        result: dict[str, Any] = {
            "source": str(item.source),
            "destination": str(item.destination),
            "category": item.category,
            "run_name": item.run_name,
            "status": "failed",
            "backup": "",
        }
        temporary: Path | None = None
        try:
            if item.category not in roots:
                raise JobError(f"Unknown completed-output category: {item.category}")
            source_root, destination_root = roots[item.category]
            source = item.source.resolve()
            destination = item.destination.resolve()
            if not source.is_file():
                raise JobError("Completed output no longer exists.")
            if not source.is_relative_to(source_root):
                raise JobError("Completed output is outside its controlled workspace.")
            if not destination.is_relative_to(destination_root):
                raise JobError("Production destination is outside the selected job folder.")

            destination.parent.mkdir(parents=True, exist_ok=True)
            if destination.exists() and filecmp.cmp(source, destination, shallow=False):
                source.unlink()
                result["status"] = "already_current"
            else:
                if destination.exists():
                    relative = destination.relative_to(destination_root)
                    backup = backup_root / item.category / relative
                    backup.parent.mkdir(parents=True, exist_ok=True)
                    if backup.exists():
                        backup = backup.with_name(
                            f"{backup.stem}_{uuid.uuid4().hex[:8]}{backup.suffix}"
                        )
                    shutil.copy2(destination, backup)
                    result["backup"] = str(backup)

                temporary = destination.with_name(
                    f".{destination.name}.{uuid.uuid4().hex}.moving"
                )
                shutil.copy2(source, temporary)
                if temporary.stat().st_size != source.stat().st_size:
                    raise JobError("Copied output size did not match its source.")
                os.replace(temporary, destination)
                temporary = None
                source.unlink()
                result["status"] = "replaced" if result["backup"] else "moved"

            _remove_empty_output_parents(source, source_root)
        except (OSError, JobError) as exc:
            result["error"] = str(exc)
        finally:
            if temporary is not None and temporary.exists():
                try:
                    temporary.unlink()
                except OSError:
                    pass
        results.append(result)

    report = {
        "schema_version": 1,
        "created_at": utc_now(),
        "user": operator_name(),
        "job_number": manifest["job"]["number"],
        "revision": manifest["job"]["revision"],
        "counts": {
            status: sum(result["status"] == status for result in results)
            for status in (
                "moved",
                "replaced",
                "already_current",
                "failed",
            )
        },
        "results": results,
    }
    report_path = Path(manifest["workspace"]["logs"]) / (
        f"completed-output-move-{stamp}-{uuid.uuid4().hex[:8]}.json"
    )
    report_path_value = ""
    try:
        _atomic_json(report, report_path)
        report_path_value = str(report_path)
    except OSError as exc:
        for result in results:
            result.setdefault("report_warning", str(exc))

    successful = {
        "moved",
        "replaced",
        "already_current",
    }
    manifest.setdefault("output_moves", []).extend(
        {
            "at": utc_now(),
            "source": result["source"],
            "destination": result["destination"],
            "category": result["category"],
            "run_name": result["run_name"],
            "status": result["status"],
            "backup": result.get("backup", ""),
        }
        for result in results
        if result["status"] in successful
    )
    record_event(
        manifest,
        "completed_outputs_moved",
        results=results,
        report=report_path_value,
    )
    for result in results:
        result["report"] = report_path_value
    return results


def parse_comparison_summary(folder: Path) -> dict[str, Any]:
    """Read the stable JSON contract, including a timestamped child run.

    Older comparison outputs remain supported through their compact CSVs.
    """
    summaries = (
        sorted(
            folder.glob("**/comparison_summary.json"),
            key=lambda path: path.stat().st_mtime,
            reverse=True,
        )
        if folder.is_dir()
        else []
    )
    if summaries:
        try:
            payload = json.loads(summaries[0].read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as exc:
            return {
                "status": "not_available",
                "errors": 0,
                "warnings": 0,
                "message": f"Comparison summary is invalid: {exc}",
            }
        if payload.get("schema_version") != 1:
            return {
                "status": "not_available",
                "errors": 0,
                "warnings": 0,
                "message": "Comparison summary uses an unsupported schema version.",
            }
        counts = payload.get("counts", {})
        reports = payload.get("reports", {})
        errors = int(counts.get("errors", 0)) + int(counts.get("missing_core", 0))
        warnings = int(counts.get("not_checked", 0)) + int(
            counts.get("source_issues", 0)
        )
        status = payload.get("outcome", "not_available")
        messages = {
            "action_required": "Action required",
            "review_recommended": "Review recommended",
            "no_discrepancies": "No discrepancies found",
        }
        return {
            "status": status,
            "errors": errors,
            "warnings": warnings,
            "message": messages.get(status, "Comparison result is unavailable"),
            "excel": reports.get("excel", ""),
            "html": reports.get("html", ""),
            "folder": reports.get("folder", str(summaries[0].parent)),
            "summary": str(summaries[0]),
        }

    run_folders = (
        sorted(
            (path.parent for path in folder.glob("**/errors_requiring_action.csv")),
            key=lambda path: path.stat().st_mtime,
            reverse=True,
        )
        if folder.is_dir()
        else []
    )
    result_folder = run_folders[0] if run_folders else folder
    action = result_folder / "errors_requiring_action.csv"
    issues = result_folder / "source_data_issues.csv"
    if not action.exists() or not issues.exists():
        return {
            "status": "not_available",
            "errors": 0,
            "warnings": 0,
            "message": "Comparison outputs are incomplete; review the process log.",
        }

    def rows(path: Path) -> int:
        with path.open("r", encoding="utf-8-sig", errors="replace") as handle:
            return max(0, sum(1 for _ in handle) - 1)

    errors, warnings = rows(action), rows(issues)
    status = (
        "action_required"
        if errors
        else "review_recommended"
        if warnings
        else "no_discrepancies"
    )
    messages = {
        "action_required": "Action required",
        "review_recommended": "Review recommended",
        "no_discrepancies": "No discrepancies found",
    }
    return {
        "status": status,
        "errors": errors,
        "warnings": warnings,
        "message": messages[status],
        "excel": str(result_folder / "production_part_comparison.xlsx"),
        "html": str(result_folder / "comparison_report.html"),
        "folder": str(result_folder),
    }


# Compatibility aliases retained for older callers.
preflight = stage_checks


def load_local_config(repo_root: Path) -> dict[str, Any]:
    """Compatibility wrapper for the former repository-local configuration."""
    return load_settings(repo_root=repo_root)
