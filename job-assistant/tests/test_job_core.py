import io
import csv
import json
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from job_core import (  # noqa: E402
    MANIFEST_VERSION,
    JobError,
    OutputMoveItem,
    acknowledge_override,
    command_bom,
    command_comparison,
    command_dxf,
    dashboard_warnings,
    discover_drawings,
    existing_working_directory,
    export_parts_list_csv,
    load_manifest,
    load_settings,
    migrate_manifest,
    move_completed_outputs,
    inferred_plate_thickness,
    parse_comparison_summary,
    plan_completed_outputs,
    prepare_dxf_workspace,
    save_manifest,
    save_settings,
    setup_job,
    stage_checks,
    suggest_job_number,
)
from job_assistant import JobAssistant  # noqa: E402


class CoreTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name) / "01234 Customer" / "SA ENGINEERING PROCESS"
        self.model = self.root / "odd MODEL name"
        self.cut = self.root / "incoming cuts"
        self.model.mkdir(parents=True)
        self.cut.mkdir()

    def tearDown(self):
        self.temp.cleanup()

    def test_plate_thickness_inference(self):
        self.assertEqual(inferred_plate_thickness(Path("250-A36")), 0.25)
        self.assertEqual(inferred_plate_thickness(Path("375-HSLA-65")), 0.375)
        self.assertIsNone(inferred_plate_thickness(Path(self.temp.name) / "notes"))

    def manifest(self):
        path = setup_job(self.root, self.model, self.cut, "01234", "Test", "A")
        return load_manifest(path), path

    def test_setup_uses_selected_paths_and_isolated_workspace(self):
        manifest, path = self.manifest()
        self.assertEqual(Path(manifest["paths"]["model_3d"]), self.model)
        self.assertEqual(path.parent.name, "_JOB_ASSISTANT")
        self.assertTrue(Path(manifest["workspace"]["staging"]).is_dir())
        self.assertEqual(suggest_job_number(self.root), "01234")

    def test_setup_never_replaces_existing_manifest_history(self):
        manifest, path = self.manifest()
        manifest["events"].append({"type": "keep_me"})
        save_manifest(manifest, path)
        with self.assertRaisesRegex(JobError, "Open it instead"):
            setup_job(self.root, self.model, self.cut, "99999", "Replace", "Z")
        self.assertEqual(load_manifest(path)["events"][-1]["type"], "keep_me")

    def test_atomic_save_leaves_no_temporary_file(self):
        manifest, path = self.manifest()
        save_manifest(manifest, path)
        self.assertEqual(
            json.loads(path.read_text())["manifest_version"], MANIFEST_VERSION
        )
        self.assertEqual(list(path.parent.glob("*.tmp")), [])

    def test_v1_manifest_migrates_missing_keys(self):
        old = {
            "manifest_version": 1,
            "root": str(self.root),
            "job": {"number": "1", "name": "x", "revision": "A"},
            "paths": {"cad_models": str(self.model), "cut_files": str(self.cut)},
            "stages": {},
            "events": [],
        }
        migrated = migrate_manifest(old)
        self.assertEqual(migrated["manifest_version"], MANIFEST_VERSION)
        self.assertIn("comparison", migrated)
        self.assertEqual(migrated["paths"]["model_3d"], str(self.model))
        self.assertEqual(migrated["events"][-1]["type"], "manifest_migrated")

    def test_removed_manual_stages_are_preserved_as_legacy_history(self):
        old = {
            "manifest_version": 3,
            "root": str(self.root),
            "job": {"number": "1", "name": "x", "revision": "A"},
            "paths": {"model_3d": str(self.model), "cut_files": str(self.cut)},
            "stages": {
                "manual_model": {
                    "status": "complete",
                    "notes": "Reviewed specialized model",
                },
                "final": {"status": "complete", "notes": "Final review"},
            },
        }

        migrated = migrate_manifest(old)

        self.assertNotIn("manual_model", migrated["stages"])
        self.assertNotIn("final", migrated["stages"])
        self.assertEqual(
            migrated["legacy_stages"]["manual_model"]["notes"],
            "Reviewed specialized model",
        )
        self.assertEqual(len(migrated["stages"]), 5)

    def test_future_manifest_has_useful_error(self):
        with self.assertRaisesRegex(JobError, "supports up to"):
            migrate_manifest({"manifest_version": 999})

    def test_incomplete_manifest_has_useful_error(self):
        with self.assertRaisesRegex(JobError, "job number is missing"):
            migrate_manifest(
                {
                    "manifest_version": MANIFEST_VERSION,
                    "root": str(self.root),
                    "job": {"revision": "A"},
                }
            )

    def test_settings_round_trip_and_defaults(self):
        target = Path(self.temp.name) / "local" / "settings.json"
        save_settings(
            {"parts_list_template": r"\\server\share\Parts List.xlsx"}, target
        )
        loaded = load_settings(target)
        self.assertEqual(
            loaded["parts_list_template"], r"\\server\share\Parts List.xlsx"
        )
        self.assertIn("solidworks_executable", loaded)
        self.assertEqual(loaded["autocad_workers"], 2)

    def test_autocad_worker_setting_is_bounded(self):
        target = Path(self.temp.name) / "local" / "settings.json"
        save_settings({"autocad_workers": 99}, target)
        self.assertEqual(load_settings(target)["autocad_workers"], 4)

    def test_drawing_discovery_defaults_and_safe_numbered_copy(self):
        dxf = self.cut / "Plate A.dxf"
        dwg = self.cut / "Shape B.DWG"
        txt = self.cut / "note.txt"
        dxf.write_text("dxf")
        dwg.write_text("dwg")
        txt.write_text("note")
        candidates = discover_drawings(self.cut)
        self.assertEqual(
            [(c.path.name, c.selected) for c in candidates],
            [("Plate A.dxf", True), ("Shape B.DWG", False)],
        )
        manifest, _ = self.manifest()
        parts_csv = self.root / "Parts List.csv"
        parts_csv.write_text("PartNumber,Quantity,Thickness,Material\nPlate A,1,1/2,A36\n")
        run = prepare_dxf_workspace(manifest, [dxf], parts_csv)
        self.assertEqual((run / "001" / dxf.name).read_text(), "dxf")
        self.assertEqual((run / "Parts List.csv").read_text(), parts_csv.read_text())
        self.assertTrue(dxf.exists())

        second_run = prepare_dxf_workspace(manifest, [dxf], parts_csv)
        self.assertNotEqual(run, second_run)
        self.assertTrue(second_run.name.startswith(run.name))

    def test_warnings_are_overridable_and_audited(self):
        manifest, _ = self.manifest()
        checks = stage_checks(manifest, "dxf")
        self.assertTrue(any(c.code == "previous_stage_incomplete" for c in checks))
        acknowledge_override(manifest, "dxf", checks)
        event = manifest["events"][-1]
        self.assertEqual(event["type"], "warning_overridden")
        self.assertTrue(event["user"])

    def test_completed_cut_files_and_plate_models_move_to_production(self):
        manifest, _ = self.manifest()
        manifest["stages"]["dxf"]["status"] = "complete"
        manifest["stages"]["plate_model"]["status"] = "complete"
        cut_source = (
            Path(manifest["workspace"]["working"])
            / "DXF Orchestrator"
            / "20260805-120000"
            / "250-A36"
            / "PLATE-1.dwg"
        )
        plate_source = (
            Path(manifest["workspace"]["staging"])
            / "SolidWorks Parts"
            / "250-A36"
            / "PLATE-1.SLDPRT"
        )
        cut_source.parent.mkdir(parents=True)
        plate_source.parent.mkdir(parents=True)
        cut_source.write_text("new drawing")
        plate_source.write_text("new model")
        existing_cut = self.cut / "250-A36" / "PLATE-1.dwg"
        existing_cut.parent.mkdir()
        existing_cut.write_text("old drawing")

        plan = plan_completed_outputs(manifest)

        self.assertEqual(len(plan), 2)
        cut_item = next(item for item in plan if item.category == "cut_file")
        plate_item = next(item for item in plan if item.category == "plate_model")
        self.assertEqual(cut_item.destination, existing_cut.resolve())
        self.assertTrue(cut_item.conflict)
        self.assertEqual(plate_item.destination, (self.model / "PLATE-1.SLDPRT").resolve())

        results = move_completed_outputs(manifest, plan)

        by_category = {result["category"]: result for result in results}
        self.assertEqual(by_category["cut_file"]["status"], "replaced")
        self.assertEqual(by_category["plate_model"]["status"], "moved")
        self.assertEqual(existing_cut.read_text(), "new drawing")
        self.assertEqual((self.model / "PLATE-1.SLDPRT").read_text(), "new model")
        self.assertEqual(
            Path(by_category["cut_file"]["backup"]).read_text(), "old drawing"
        )
        self.assertFalse(cut_source.exists())
        self.assertFalse(plate_source.exists())
        self.assertTrue(Path(results[0]["report"]).is_file())
        self.assertEqual(len(manifest["output_moves"]), 2)

    def test_completed_output_with_identical_destination_removes_staged_copy(self):
        manifest, _ = self.manifest()
        source = (
            Path(manifest["workspace"]["staging"])
            / "SolidWorks Parts"
            / "250-A36"
            / "P1.SLDPRT"
        )
        source.parent.mkdir(parents=True)
        source.write_text("same")
        destination = self.model / source.name
        destination.write_text("same")
        manifest["stages"]["plate_model"]["status"] = "complete"

        result = move_completed_outputs(
            manifest,
            [OutputMoveItem(source, destination, "plate_model", "250-A36", True)],
        )[0]

        self.assertEqual(result["status"], "already_current")
        self.assertFalse(source.exists())
        self.assertFalse(result["backup"])

    def test_incomplete_stages_are_not_planned_for_output_moves(self):
        manifest, _ = self.manifest()
        source = (
            Path(manifest["workspace"]["working"])
            / "DXF Orchestrator"
            / "run"
            / "250-A36"
            / "P1.dwg"
        )
        source.parent.mkdir(parents=True)
        source.write_text("drawing")
        self.assertEqual(plan_completed_outputs(manifest), [])

    def test_comparison_summary(self):
        reports = Path(self.temp.name) / "reports"
        reports.mkdir()
        (reports / "errors_requiring_action.csv").write_text("part,status\nP1,error\n")
        (reports / "source_data_issues.csv").write_text("issue\nwarning\n")
        summary = parse_comparison_summary(reports)
        self.assertEqual(summary["status"], "action_required")
        self.assertEqual((summary["errors"], summary["warnings"]), (1, 1))

    def test_dashboard_warnings_summarize_actionable_comparison(self):
        manifest, _ = self.manifest()
        manifest["comparison"] = {"status": "action_required", "errors": 4}
        self.assertIn(
            "Production comparison has 4 actionable discrepancy item(s).",
            dashboard_warnings(manifest),
        )

    def test_versioned_comparison_summary_in_timestamped_folder(self):
        reports = Path(self.temp.name) / "reports"
        run = reports / "Part_Comparison_2026-01-02_030405"
        run.mkdir(parents=True)
        payload = {
            "schema_version": 1,
            "outcome": "action_required",
            "counts": {
                "errors": 2,
                "missing_core": 1,
                "not_checked": 3,
                "source_issues": 4,
            },
            "reports": {
                "excel": str(run / "production_part_comparison.xlsx"),
                "html": str(run / "comparison_report.html"),
                "folder": str(run),
            },
        }
        (run / "comparison_summary.json").write_text(json.dumps(payload))
        summary = parse_comparison_summary(reports)
        self.assertEqual(summary["status"], "action_required")
        self.assertEqual((summary["errors"], summary["warnings"]), (3, 7))
        self.assertEqual(Path(summary["folder"]), run)

    def test_cad_batch_source_uses_runtime_paths_not_job_constants(self):
        config = (
            Path(__file__).resolve().parents[2]
            / "solidworks/cad-batch-converter/Config.bas"
        ).read_text(encoding="utf-8")
        self.assertIn("Environ$(environmentName)", config)
        self.assertIn("GetSetting(SETTINGS_APP, SETTINGS_SECTION", config)
        self.assertIn('"MACROS_EXTRUDE_DEPTH_METERS"', config)
        self.assertIn('"ExtrudeDepthMeters"', config)
        self.assertNotIn("Public Const EXTRUDE_DEPTH_METERS", config)
        self.assertLess(
            config.index('"ExtrudeDepthMeters"'),
            config.index('Environ$("MACROS_EXTRUDE_DEPTH_METERS")'),
        )
        self.assertNotIn("U:\\Engineering\\CAD Services\\Current Jobs", config)

    def test_powershell_bevel_path_is_marked_and_headless(self):
        orchestrator = (
            Path(__file__).resolve().parents[2]
            / "autocad/dxf-orchestrator/Master_Orchestrator.ps1"
        ).read_text(encoding="utf-8-sig")
        self.assertIn('$workingName = "${workingName}(B)"', orchestrator)
        self.assertIn("Start-Process $AcadConsolePath", orchestrator)
        self.assertNotIn("$AcadGuiPath", orchestrator)
        self.assertNotIn("$finishLisp", orchestrator)
        self.assertNotIn("SPCFINISH", orchestrator)

    def test_external_commands_are_argument_lists_with_spaces(self):
        repo = Path(r"Z:\Shared Macros")
        bom = command_bom(
            "python.exe",
            repo,
            Path("source copy.xlsx"),
            Path("out.xlsx"),
            Path("template.xlsx"),
        )
        self.assertEqual(bom[-3:], ["source copy.xlsx", "out.xlsx", "template.xlsx"])
        self.assertIn("--gui", bom)
        compare = command_comparison(
            "python.exe",
            repo,
            Path("Nest Data"),
            Path("parts.csv"),
            Path("sw.csv"),
            Path("Reports"),
        )
        self.assertIn("Nest Data", compare)

        dxf = command_dxf(
            "python.exe",
            repo,
            Path(r"C:\Job Workspace"),
            Path(r"C:\Job Workspace\Parts List.csv"),
            autocad_console=Path(r"C:\CAD Suite\accoreconsole.exe"),
        )
        self.assertEqual(
            dxf[-2:],
            [
                "--acad-console-path",
                r"C:\CAD Suite\accoreconsole.exe",
            ],
        )
        self.assertNotIn("--acad-gui-path", dxf)
        self.assertEqual(dxf[0], "python.exe")
        self.assertIn("-u", dxf)
        self.assertIn("Master_Orchestrator.py", " ".join(dxf))
        self.assertIn("--workers", dxf)

        packaged = command_bom(
            "ignored-python.exe",
            repo,
            Path("source.xlsx"),
            Path("out.xlsx"),
            Path("template.xlsx"),
            Path("Engineering BOM Converter.exe"),
        )
        self.assertEqual(packaged[0], "Engineering BOM Converter.exe")
        self.assertNotIn("bom_converter.py", " ".join(packaged))

    def test_missing_working_directory_is_not_passed_to_subprocess(self):
        missing = self.root / "disconnected macros share"
        self.assertIsNone(existing_working_directory(missing))
        self.assertEqual(existing_working_directory(self.root), self.root)

    def test_generated_parts_list_exports_orchestrator_csv(self):
        workbook_path = self.root / "123_PARTS_LIST.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["title"])
        sheet.append(
            [
                "PART NUMBER",
                "DESCRIPTION",
                "QUANTITY",
                "THICKNESS/SHAPE",
                "MATERIAL TYPE",
            ]
        )
        sheet.append(["PLATE-1", "PLATE", 3, "1/2", "A36"])
        sheet.append(["ANGLE-1", "ANGLE", 2, "L3x3x1/4", "A36"])
        workbook.save(workbook_path)
        output = export_parts_list_csv(workbook_path, self.root / "Parts List.csv")
        with output.open(encoding="utf-8-sig") as handle:
            rows = list(csv.DictReader(handle))
        self.assertEqual(
            rows,
            [{
                "PartNumber": "PLATE-1",
                "Quantity": "3",
                "Thickness": "1/2",
                "Material": "A36",
            }],
        )

    def test_parts_list_export_accepts_documented_aliases_and_plate_labels(self):
        workbook_path = self.root / "aliased.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(
            [
                "part-number",
                "description",
                "total quantity",
                "thickness shape",
                "material",
            ]
        )
        sheet.append(["P-PL", " pl ", None, "3/8", "A572"])
        sheet.append(["P-PLATE", "Plate", 4, "1/2", "A36"])
        sheet.append(["P-PIPE", "PIPE", 9, "6 SCH 40", "A106"])
        workbook.save(workbook_path)

        destination = self.root / "nested" / "parts.csv"
        output = export_parts_list_csv(workbook_path, destination)

        with output.open(encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            self.assertEqual(
                reader.fieldnames,
                ["PartNumber", "Quantity", "Thickness", "Material"],
            )
            self.assertEqual(
                list(reader),
                [
                    {
                        "PartNumber": "P-PL",
                        "Quantity": "1",
                        "Thickness": "3/8",
                        "Material": "A572",
                    },
                    {
                        "PartNumber": "P-PLATE",
                        "Quantity": "4",
                        "Thickness": "1/2",
                        "Material": "A36",
                    },
                ],
            )

    def test_parts_list_export_rejects_workbook_without_description(self):
        workbook_path = self.root / "missing-description.xlsx"
        workbook = openpyxl.Workbook()
        workbook.active.append(
            ["PART NUMBER", "QUANTITY", "THICKNESS/SHAPE", "MATERIAL TYPE"]
        )
        workbook.save(workbook_path)
        destination = self.root / "Parts List.csv"

        with self.assertRaisesRegex(JobError, "DESCRIPTION"):
            export_parts_list_csv(workbook_path, destination)

        self.assertFalse(destination.exists())

    def test_completing_bom_stage_exports_and_records_plate_csv(self):
        manifest, manifest_path = self.manifest()
        workbook_path = self.root / "reviewed-parts.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(
            [
                "PART NUMBER",
                "DESCRIPTION",
                "QUANTITY",
                "THICKNESS/SHAPE",
                "MATERIAL TYPE",
            ]
        )
        sheet.append(["PLATE-10", "PLATE", 2, "1/4", "A36"])
        sheet.append(["CHANNEL-10", "CHANNEL", 5, "C6x8.2", "A36"])
        workbook.save(workbook_path)
        manifest["stages"]["bom"].update(
            status="needs_review", parts_list_workbook=str(workbook_path)
        )

        assistant = object.__new__(JobAssistant)
        assistant.manifest = manifest
        assistant.manifest_path = manifest_path
        assistant.require_job = lambda: None
        assistant.selected_stage = lambda: "bom"
        assistant.handle = lambda operation: operation()
        assistant.refresh = lambda: None

        with patch(
            "job_assistant.simpledialog.askstring", return_value="Checked plates"
        ):
            assistant.finish_stage()

        saved = load_manifest(manifest_path)
        csv_path = Path(saved["workspace"]["source_copies"]) / "Parts List.csv"
        self.assertEqual(saved["stages"]["bom"]["status"], "complete")
        self.assertEqual(
            {item["name"] for item in saved["stages"]["bom"]["artifacts"]},
            {"reviewed-parts.xlsx", "Parts List.csv"},
        )
        with csv_path.open(encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.DictReader(handle))
        self.assertEqual([row["PartNumber"] for row in rows], ["PLATE-10"])

    def test_completing_bom_stage_recovers_moved_workbook_with_prompt(self):
        manifest, manifest_path = self.manifest()
        workbook_path = self.root / "moved-reviewed-parts.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(
            ["PART NUMBER", "DESCRIPTION", "THICKNESS/SHAPE", "MATERIAL TYPE"]
        )
        sheet.append(["PLATE-20", "PL", "3/4", "A572"])
        workbook.save(workbook_path)
        manifest["stages"]["bom"].update(
            status="needs_review",
            parts_list_workbook=str(self.root / "old-location.xlsx"),
        )

        assistant = object.__new__(JobAssistant)
        assistant.manifest = manifest
        assistant.manifest_path = manifest_path
        assistant.require_job = lambda: None
        assistant.selected_stage = lambda: "bom"
        assistant.handle = lambda operation: operation()
        assistant.refresh = lambda: None

        with (
            patch("job_assistant.simpledialog.askstring", return_value="Reviewed"),
            patch(
                "job_assistant.filedialog.askopenfilename",
                return_value=str(workbook_path),
            ) as choose_workbook,
        ):
            assistant.finish_stage()

        choose_workbook.assert_called_once()
        saved = load_manifest(manifest_path)
        self.assertEqual(
            saved["stages"]["bom"]["parts_list_workbook"], str(workbook_path)
        )
        self.assertEqual(saved["stages"]["bom"]["status"], "complete")

    def test_finished_dxf_process_updates_the_job_that_launched_it(self):
        first_manifest, first_path = self.manifest()
        other_root = Path(self.temp.name) / "99999 Other"
        other_model = other_root / "model"
        other_cut = other_root / "cuts"
        other_model.mkdir(parents=True)
        other_cut.mkdir()
        other_path = setup_job(
            other_root, other_model, other_cut, "99999", "Other", "A"
        )
        other_manifest = load_manifest(other_path)

        assistant = object.__new__(JobAssistant)
        assistant.manifest = other_manifest
        assistant.manifest_path = other_path
        assistant.refresh = lambda: self.fail("the unrelated open job was refreshed")
        log = Path(first_manifest["workspace"]["logs"]) / "dxf-test.log"
        log.parent.mkdir(parents=True, exist_ok=True)
        log.write_text("started\n")

        class FinishedProcess:
            pid = 42

            @staticmethod
            def poll():
                return 0

        handle = io.StringIO()
        assistant.post_background_notice = lambda title, message, **kwargs: setattr(
            assistant, "completion_notice", (title, message, kwargs)
        )
        with patch("job_assistant.messagebox.showinfo") as modal_notification:
            assistant._poll_dxf_process(
                FinishedProcess(),
                handle,
                log,
                Path(first_manifest["workspace"]["working"]),
                first_manifest,
                first_path,
            )

        saved_first = load_manifest(first_path)
        saved_other = load_manifest(other_path)
        self.assertEqual(saved_first["stages"]["dxf"]["status"], "needs_review")
        self.assertEqual(saved_other["stages"]["dxf"]["status"], "not_started")
        self.assertIn("01234", assistant.completion_notice[1])
        modal_notification.assert_not_called()

    def test_finished_comparison_posts_nonmodal_dashboard_notice(self):
        manifest, manifest_path = self.manifest()
        output = Path(manifest["workspace"]["reports"]) / "comparison-A"
        run = output / "Part_Comparison_2026-08-05_120000"
        run.mkdir(parents=True)
        (run / "comparison_summary.json").write_text(
            json.dumps(
                {
                    "schema_version": 1,
                    "outcome": "review_recommended",
                    "counts": {
                        "errors": 0,
                        "missing_core": 0,
                        "not_checked": 1,
                        "source_issues": 0,
                    },
                    "reports": {
                        "excel": str(run / "report.xlsx"),
                        "html": str(run / "report.html"),
                        "folder": str(run),
                    },
                }
            ),
            encoding="utf-8",
        )
        log = Path(manifest["workspace"]["logs"]) / "comparison-test.log"
        log.write_text("started\n", encoding="utf-8")

        class FinishedProcess:
            pid = 43

            @staticmethod
            def poll():
                return 0

        assistant = object.__new__(JobAssistant)
        assistant.manifest = manifest
        assistant.manifest_path = manifest_path
        assistant.running_processes = {43: {"stage": "comparison"}}
        assistant.update_running_summary = lambda: None
        assistant.refresh = lambda: None
        assistant.show_stage = lambda: None
        assistant.post_background_notice = (
            lambda title, message, **kwargs: setattr(
                assistant, "completion_notice", (title, message, kwargs)
            )
        )

        with (
            patch("job_assistant.messagebox.showinfo") as show_info,
            patch("job_assistant.messagebox.showerror") as show_error,
        ):
            assistant._poll_comparison_process(
                FinishedProcess(),
                io.StringIO(),
                log,
                output,
                manifest,
                manifest_path,
            )

        saved = load_manifest(manifest_path)
        self.assertEqual(saved["stages"]["comparison"]["status"], "needs_review")
        self.assertIn("Production comparison finished", assistant.completion_notice[0])
        show_info.assert_not_called()
        show_error.assert_not_called()


if __name__ == "__main__":
    unittest.main()
