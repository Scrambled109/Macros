from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from epl_converter import ConversionError, convert_epls


HEADERS = [
    "PS LVL",
    "PART NUMBER",
    "QTY",
    "UM",
    "PART DESCRIPTION",
    "PART TYPE",
    "PART MODIFIER",
    "MF PART NO.",
    "MF NOUN NAME",
    "MF TYPE",
    "MF MODIFIER",
    "MF CAGE/SPECIFICATION",
    "MF DESCRIPTION",
]


def make_epl(path: Path, rows: list[list[object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "EPL"
    ws.append(["ENGINEERING PARTS LIST"])
    ws.append(HEADERS)
    for row in rows:
        ws.append(row)
    wb.save(path)


def make_bop(
    path: Path,
    rows: list[tuple[str, str, str, str, object]],
    hull_no: str = "828",
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "BOP"
    ws["A1"] = f"{hull_no}_TEST BOP"
    ws.append([])
    for _ in range(7):
        ws.append([])
    ws.append(
        [
            "BOM Line",
            None,
            "DYPN",
            None,
            None,
            "eb4_ref_comp_loa_id",
            "CSA",
            "Hold-Up",
            "LT/MFG/PC Cd",
            "Revision",
            "MF Part Number",
            "Quantity",
        ]
    )
    for loa, part_no, mdlprt, mf_part_no, quantity in rows:
        ws.append(
            [
                f"{mdlprt}/A;1-TEST PART x {quantity}",
                mdlprt,
                part_no,
                None,
                None,
                loa,
                "N",
                None,
                None,
                "A",
                mf_part_no,
                quantity,
            ]
        )
    wb.save(path)


class ConverterTests(unittest.TestCase):
    def setUp(self) -> None:
        self.tmp = tempfile.TemporaryDirectory()
        self.root = Path(self.tmp.name)

    def tearDown(self) -> None:
        self.tmp.cleanup()

    def test_conversion_classification_materials_and_metadata(self) -> None:
        source = self.root / "LOA012095_B-EPL.xlsx"
        bop = self.root / "Test-BOP.xlsx"
        make_epl(
            source,
            [
                ["1", "R5610102-A1", "1", "PC", "ASSEMBLY", "", "", "", "", "", "", "", ""],
                [
                    "2",
                    "R5610102-1",
                    "2",
                    "PC",
                    "PLATE",
                    "HIGH STRENGTH",
                    "",
                    "EB218001124",
                    "PLATE",
                    "HIGH STRENGTH",
                    "",
                    "81349 / MIL-S-22698_CL-P,GR-EH-36T",
                    "STEEL 0.500 THK",
                ],
                [
                    "2",
                    "R5610102-2",
                    "1",
                    "PC",
                    "HOUSING",
                    "",
                    "",
                    "211006856",
                    "HOUSING",
                    "STUFFING TUBE",
                    "",
                    "QQ-S-763 CL-316L",
                    "CRES 316L 4.500IDX6.750OD",
                ],
                [
                    "2",
                    "R5610102-3",
                    "1",
                    "PC",
                    "PIG",
                    "",
                    "",
                    "999",
                    "PIG",
                    "LEAD",
                    "",
                    "QQ-L-171",
                    "LEAD 8.000LG",
                ],
            ],
        )
        make_bop(
            bop,
            [
                ("LOA012095", "R5610102-1", "MDLPRT1", "EB218001124", 2),
                ("LOA012095", "R5610102-2", "MDLPRT2", "211006856", 1),
                ("LOA012095", "R5610102-3", "MDLPRT3", "999", 1),
            ],
        )
        metadata = {
            "R5610102-1": {
                "mdlprt": "MDLPRT1",
                "width": '48"',
                "length": "8' - 0\"",
                "beveled": True,
                "other_info": "CHECK",
            },
            "R5610102-2": {
                "mdlprt": "MDLPRT2",
                "length": '3 1/4"',
                "beveled": False,
            },
        }
        result = convert_epls(
            [source], self.root, "Test Output", metadata, bop_paths=[bop]
        )
        self.assertEqual(result.plate_count, 1)
        self.assertEqual(result.shape_count, 1)
        self.assertEqual(result.assembly_count, 1)
        self.assertEqual(result.unclassified_count, 1)

        plate_ws = load_workbook(result.plates_path, data_only=True).active
        self.assertEqual(plate_ws.title, "LOA012095 - R5610102")
        self.assertEqual(plate_ws["I4"].value, "MIL-S-22698_EH-36T")
        self.assertEqual(plate_ws["J4"].value, "0.500 THK")
        self.assertEqual(plate_ws["S4"].value, "CHECK; BEVELED")
        self.assertEqual(plate_ws["H4"].number_format, "@")
        self.assertEqual(plate_ws["C4"].value, "828")
        self.assertEqual(plate_ws.auto_filter.ref, "B3:S4")
        self.assertEqual(plate_ws["B2"].fill.fgColor.rgb, "000070C0")

        shape_ws = load_workbook(result.shapes_path, data_only=True).active
        self.assertEqual(shape_ws["F3"].value, "STUFFING TUBE")
        self.assertEqual(shape_ws["I3"].value, "CRES 316L")
        self.assertEqual(shape_ws["J3"].value, "4.500IDX6.750OD")
        self.assertAlmostEqual(shape_ws["L3"].value, 3.25)
        self.assertNotIn("BEVELED", shape_ws["P3"].value or "")

    def test_duplicate_loa_gets_unique_sheet_names_and_rows_remain(self) -> None:
        source = self.root / "LOA123456_A-EPL.xlsx"
        bop = self.root / "Duplicates-BOP.xlsx"
        row = [
            "2",
            "ABC-1",
            "1",
            "PC",
            "PLATE",
            "",
            "",
            "MF1",
            "PLATE",
            "HY-80",
            "",
            "T9074 HY-80",
            "STEEL 1.000 THK",
        ]
        make_epl(source, [row, row])
        make_bop(bop, [("LOA123456", "ABC-1", "MDLPRT1", "MF1", 1)])
        result = convert_epls(
            [source, source], self.root, "Duplicates", bop_paths=[bop]
        )
        wb = load_workbook(result.plates_path, read_only=True)
        self.assertEqual(wb.sheetnames, ["LOA123456 - ABC", "LOA123456 - ABC (2)"])
        self.assertEqual(result.plate_count, 4)
        self.assertTrue(any(i.code == "DUPLICATE_PART_NUMBER" for i in result.issues))

    def test_unknown_material_is_preserved_and_reported(self) -> None:
        source = self.root / "LOA000001-EPL.xlsx"
        bop = self.root / "Unknown-BOP.xlsx"
        make_epl(
            source,
            [
                [
                    "2",
                    "TEST-1",
                    "1",
                    "PC",
                    "SHEET",
                    "",
                    "",
                    "MF1",
                    "SHEET",
                    "MYSTERY ALLOY",
                    "",
                    "SPEC-XYZ",
                    "MYSTERY ALLOY 0.125 THK",
                ]
            ],
        )
        make_bop(bop, [("LOA000001", "TEST-1", "MDLPRT1", "MF1", 1)])
        result = convert_epls([source], self.root, "Unknown", bop_paths=[bop])
        ws = load_workbook(result.plates_path, data_only=True).active
        self.assertEqual(ws["I4"].value, "MYSTERY ALLOY")
        self.assertTrue(any(i.code == "UNMAPPED_MATERIAL" for i in result.issues))
        report = json.loads(result.report_json_path.read_text(encoding="utf-8"))
        self.assertEqual(report["summary"]["plates_exported"], 1)

    def test_missing_epl_sheet_is_clear_error(self) -> None:
        source = self.root / "LOA000002-EPL.xlsx"
        bop = self.root / "Bad-BOP.xlsx"
        wb = Workbook()
        wb.active.title = "Not EPL"
        wb.save(source)
        make_bop(bop, [("LOA000002", "TEST-1", "MDLPRT1", "MF1", 1)])
        with self.assertRaisesRegex(ConversionError, "sheet named EPL"):
            convert_epls([source], self.root, "Bad", bop_paths=[bop])

    def test_31_character_sheet_limit(self) -> None:
        source = self.root / "LOA12345678901234567890-EPL.xlsx"
        bop = self.root / "Long-BOP.xlsx"
        make_epl(
            source,
            [
                [
                    "2",
                    "VERYLONGDRAWINGIDENTIFIER-1",
                    "1",
                    "PC",
                    "PLATE",
                    "",
                    "",
                    "MF",
                    "PLATE",
                    "HY-80",
                    "",
                    "T9074 HY-80",
                    "STEEL 0.500 THK",
                ]
            ],
        )
        make_bop(
            bop,
            [
                (
                    "LOA12345678901234567890",
                    "VERYLONGDRAWINGIDENTIFIER-1",
                    "MDLPRT1",
                    "MF",
                    1,
                )
            ],
        )
        result = convert_epls([source], self.root, "Long", bop_paths=[bop])
        wb = load_workbook(result.plates_path, read_only=True)
        self.assertLessEqual(len(wb.sheetnames[0]), 31)

    def test_bop_is_required_and_gates_scope(self) -> None:
        source = self.root / "LOA000003-EPL.xlsx"
        make_epl(
            source,
            [
                ["2", "SCOPE-1", 1, "PC", "PLATE", "", "", "MF1", "PLATE", "HY-80", "", "T9074 HY-80", "STEEL 0.5 THK"],
                ["2", "SCOPE-2", 1, "PC", "PLATE", "", "", "MF2", "PLATE", "HY-80", "", "T9074 HY-80", "STEEL 0.5 THK"],
            ],
        )
        with self.assertRaisesRegex(ConversionError, "BOP"):
            convert_epls([source], self.root, "No BOP")
        bop = self.root / "Scope-BOP.xlsx"
        make_bop(bop, [("LOA000003", "SCOPE-2", "MDLPRT2", "MF2", 1)])
        result = convert_epls([source], self.root, "Scoped", bop_paths=[bop])
        self.assertEqual(result.plate_count, 1)
        self.assertEqual(result.inputs[0].out_of_scope_rows, 1)
        ws = load_workbook(result.plates_path, data_only=True).active
        self.assertEqual(ws["H4"].value, "SCOPE-2")

    def test_plate_template_layout_is_preserved_without_old_process_marks(self) -> None:
        source = self.root / "LOA000004-EPL.xlsx"
        bop = self.root / "Template-BOP.xlsx"
        template = self.root / "Plate-Template.xlsx"
        make_epl(
            source,
            [
                [
                    "2",
                    "TEMPLATE-1",
                    1,
                    "PC",
                    "PLATE",
                    "",
                    "",
                    "MF1",
                    "PLATE",
                    "HY-80",
                    "",
                    "T9074 HY-80",
                    "STEEL 0.5 THK",
                ]
            ],
        )
        make_bop(
            bop,
            [("LOA000004", "TEMPLATE-1", "MDLPRT1", "MF1", 1)],
        )
        wb = Workbook()
        ws = wb.active
        ws["B2"] = "ITEM NO."
        ws["B2"].fill = PatternFill("solid", fgColor="0070C0")
        ws["K2"] = "PROCESS"
        ws["K3"] = "STR CUT"
        ws["L3"] = "BEVEL (BT)"
        ws["L4"].fill = PatternFill("solid", fgColor="00FF00")
        ws.column_dimensions["H"].width = 27.5
        ws.row_dimensions[4].height = 21
        wb.save(template)

        result = convert_epls(
            [source],
            self.root,
            "Templated",
            bop_paths=[bop],
            plate_template_path=template,
        )
        output = load_workbook(result.plates_path)
        out_ws = output.active
        self.assertEqual(out_ws["B2"].value, "ITEM NO.")
        self.assertEqual(out_ws["B2"].fill.fgColor.rgb, "000070C0")
        self.assertEqual(out_ws.column_dimensions["H"].width, 27.5)
        self.assertEqual(out_ws.row_dimensions[4].height, 21)
        self.assertIsNone(out_ws["L4"].value)
        self.assertIsNone(out_ws["L4"].fill.fill_type)


if __name__ == "__main__":
    unittest.main()
