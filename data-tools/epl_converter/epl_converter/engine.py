from __future__ import annotations

import csv
import copy
import json
import os
import re
import tempfile
from collections import Counter
from fractions import Fraction
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .config import load_config
from .models import BOPEntry, ConversionResult, EPLResult, Issue, PartRecord


class ConversionError(RuntimeError):
    """A novice-friendly conversion error safe to show in the UI or CLI."""


HEADER_ALIASES = {
    "part_no": {"PARTNUMBER", "PARTNO", "DYPN"},
    "quantity": {"QTY", "QUANTITY"},
    "part_description": {"PARTDESCRIPTION", "DESCRIPTION"},
    "part_type": {"PARTTYPE"},
    "part_modifier": {"PARTMODIFIER"},
    "mf_part_no": {"MFPARTNO", "MFPARTNUMBER"},
    "mf_noun_name": {"MFNOUNNAME", "MFNOUN"},
    "mf_type": {"MFTYPE"},
    "mf_modifier": {"MFMODIFIER"},
    "material_specification": {
        "MFCAGESPECIFICATION",
        "MFSPECIFICATION",
        "MATERIALSPECIFICATION",
        "SPECIFICATION",
    },
    "material_description": {"MFDESCRIPTION", "MATERIALDESCRIPTION"},
    "thickness": {"THICKNESS", "THK"},
    "width": {"WIDTH"},
    "length": {"LENGTH", "LG"},
    "unit_weight": {"UNITWEIGHT", "UNITWEIGHTLBSFT", "WEIGHTPERFOOT"},
}

BOP_HEADER_ALIASES = {
    "bom_line": {"BOMLINE"},
    "part_no": {"DYPN", "PARTNO", "PARTNUMBER"},
    "loa": {"EB4REFCOMPLOAID", "LOA"},
    "mf_part_no": {"MFPARTNUMBER", "MFPARTNO"},
    "quantity": {"QUANTITY", "QTY"},
    "revision": {"REVISION", "REV"},
}

PLATE_TOP_HEADERS = [
    "ITEM NO.",
    "HULL NO.",
    "LOA",
    "M/F PART NO.",
    "QUANTITY",
    "DESCRIPTION",
    "PART NO.",
    "MATERIAL GRADE",
    "THICKNESS",
    "PROCESS",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "REMARKS",
]

PLATE_PROCESS_HEADERS = [
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "STR CUT",
    "BEVEL (BT)",
    "BEVEL (MAN)",
    "MACHINED",
    "ROLLED (I/H)",
    "ROLLED (O/S)",
    "FORMED (I/H)",
    "FORMED (O/S)",
    "",
]

SHAPE_HEADERS = [
    "Item #",
    "LOA",
    "MF Part No.",
    "Quantity",
    "Description",
    "MDLPRT",
    "Part No.",
    "Material Grade",
    "Size/Thickness",
    "Length",
    "Length (in.)",
    "Unit Weight (lbs/ft.)",
    "Total Weight (lbs.)",
    "STATUS",
    "other info",
]

ISSUE_HEADERS = [
    "Severity",
    "Code",
    "Source File",
    "Source Row",
    "Part No.",
    "Value",
    "Message",
]

THIN_GRAY = Side(style="thin", color="BFBFBF")
HEADER_FILL = PatternFill("solid", fgColor="000000")
GRAY_FILL = PatternFill("solid", fgColor="D9D9D9")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
GREEN_FILL = PatternFill("solid", fgColor="00B050")
AMBER_FILL = PatternFill("solid", fgColor="FFD966")
RED_FILL = PatternFill("solid", fgColor="F4CCCC")
HEADER_FONT = Font(name="Arial Narrow", size=12, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial Narrow", size=12, color="000000")
BODY_BORDER = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _header_key(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]+", "", _text(value).upper())


def _safe_output_base(value: str) -> str:
    value = _text(value)
    if not value:
        raise ConversionError("Enter an output name.")
    if any(ch in value for ch in '<>:"/\\|?*'):
        raise ConversionError('The output name cannot contain < > : " / \\ | ? *.')
    return value.rstrip(". ")


def _extract_loa(path: Path) -> str:
    match = re.search(r"(?i)(LOA\d+)", path.stem)
    if not match:
        raise ConversionError(
            f'Could not find an LOA number in "{path.name}". '
            "Expected a filename like LOA012095_B-EPL.xlsx."
        )
    return match.group(1).upper()


def _find_header_row(ws: Any) -> tuple[int, dict[str, int]]:
    for row_no in range(1, min(ws.max_row, 30) + 1):
        normalized = [_header_key(ws.cell(row_no, col).value) for col in range(1, ws.max_column + 1)]
        resolved: dict[str, int] = {}
        for canonical, aliases in HEADER_ALIASES.items():
            for index, value in enumerate(normalized, start=1):
                if value in aliases:
                    resolved[canonical] = index
                    break
        if {"part_no", "part_description"}.issubset(resolved):
            return row_no, resolved
    raise ConversionError(
        f'The "EPL" sheet in "{ws.parent.properties.title or "this workbook"}" '
        "does not contain recognizable PART NUMBER and PART DESCRIPTION headers."
    )


def _cell(ws: Any, row: int, columns: Mapping[str, int], field: str) -> Any:
    col = columns.get(field)
    return ws.cell(row, col).value if col else None


def _row_value(values: tuple[Any, ...], columns: Mapping[str, int], field: str) -> Any:
    col = columns.get(field)
    return values[col - 1] if col and col <= len(values) else None


def _drawing_identifier(part_numbers: Iterable[str]) -> str:
    values = [value for value in part_numbers if value]
    if not values:
        return ""
    prefixes = []
    for value in values:
        match = re.match(r"^(.+?)[-_](?:[A-Z]*\d+[A-Z]*)$", value.strip(), re.IGNORECASE)
        prefixes.append(match.group(1) if match else value.strip())
    most_common, count = Counter(prefixes).most_common(1)[0]
    if count == len(prefixes):
        return most_common
    common = os.path.commonprefix(values).rstrip("-_. ")
    return common or most_common


def _translate_material(
    spec: str, description: str, part_type: str, config: Mapping[str, Any]
) -> tuple[str, bool]:
    combined = " ".join([spec, description, part_type]).upper().replace("\n", " ")
    for rule in config.get("material_translations", []):
        for pattern in rule.get("patterns", []):
            if re.search(pattern, combined, flags=re.IGNORECASE | re.DOTALL):
                return _text(rule.get("name")), True
    preserved = part_type or description.splitlines()[0] if description else part_type
    return _text(preserved), False


def _parse_number(value: str) -> float | None:
    value = value.strip().replace('"', "")
    if not value:
        return None
    try:
        if " " in value and "/" in value:
            whole, fraction = value.split(None, 1)
            return float(whole) + float(Fraction(fraction))
        if "/" in value:
            return float(Fraction(value))
        return float(value)
    except (ValueError, ZeroDivisionError):
        return None


def _length_to_inches(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = _text(value).upper().replace("–", "-").replace("—", "-")
    feet_match = re.search(r"(\d+(?:\.\d+)?)\s*'", text)
    inches_match = re.search(r"(?:(\d+\s+\d+/\d+|\d+/\d+|\d+(?:\.\d+)?))\s*(?:\"|IN\b)", text)
    if feet_match:
        feet = float(feet_match.group(1))
        inches = _parse_number(inches_match.group(1)) if inches_match else 0.0
        return round(feet * 12 + (inches or 0.0), 6)
    if inches_match:
        return _parse_number(inches_match.group(1))
    lg_match = re.search(r"(\d+(?:\.\d+)?)\s*(?:LG|LENGTH)\b", text)
    if lg_match:
        return float(lg_match.group(1))
    return _parse_number(text)


def _format_inch(value: str) -> str:
    return f'{value}"'


def _parse_explicit_dimensions(description: str) -> dict[str, Any]:
    result: dict[str, Any] = {}
    clean = re.sub(r"\s+", " ", description.upper()).strip()
    thk = re.search(r"(\d+\s+\d+/\d+|\d+/\d+|\d+(?:\.\d+)?)\s*(?:THK|THICK(?:NESS)?|T)\b", clean)
    if thk:
        result["thickness"] = f"{thk.group(1)} THK"

    l_w_t = re.search(
        r"(\d+(?:\.\d+)?)\s*L(?:G)?\s*X\s*(\d+(?:\.\d+)?)\s*W\s*X\s*"
        r"(\d+(?:\.\d+)?)\s*(?:THK|T)\b",
        clean,
    )
    if l_w_t:
        result["length"] = _format_inch(l_w_t.group(1))
        result["width"] = _format_inch(l_w_t.group(2))
        result["thickness"] = f"{l_w_t.group(3)} THK"
    else:
        t_w_l = re.search(
            r"(\d+(?:\.\d+)?)\s*(?:THK|T)\s*X\s*(\d+(?:\.\d+)?)\s*"
            r"(?:W\s*)?X\s*(\d+(?:\.\d+)?)\s*(?:L|LG)?\b",
            clean,
        )
        if t_w_l:
            result["thickness"] = f"{t_w_l.group(1)} THK"
            result["width"] = _format_inch(t_w_l.group(2))
            result["length"] = _format_inch(t_w_l.group(3))

    if "length" not in result:
        lg = re.search(r"(\d+(?:\.\d+)?)\s*(?:LG|LENGTH)\b", clean)
        if lg:
            result["length"] = _format_inch(lg.group(1))

    size = re.search(
        r"((?:\d+(?:\.\d+)?\s*(?:ID|OD|DIA))(?:\s*X\s*\d+(?:\.\d+)?\s*(?:ID|OD|DIA))*)",
        clean,
    )
    if size:
        result["size_thickness"] = size.group(1).replace(" ", "")
    return result


def _classify(
    part_description: str, mf_noun: str, mf_type: str, config: Mapping[str, Any]
) -> str:
    values = [_text(part_description).upper(), _text(mf_noun).upper(), _text(mf_type).upper()]
    plate_words = {_text(x).upper() for x in config.get("plate_keywords", [])}
    shape_words = {_text(x).upper() for x in config.get("shape_keywords", [])}
    if any(any(word == value or word in value.split() for word in plate_words) for value in values):
        return "plate"
    if any(any(word == value or word in value for word in shape_words) for value in values):
        return "shape"
    return ""


def _display_description(part_description: str, mf_noun: str, mf_type: str) -> str:
    mf_type_clean = _text(mf_type)
    generic_material_types = {
        "HIGH STRENGTH",
        "HY-80",
        "LEAD",
        "STEEL",
        "RUBBER",
        "CRES 316L",
    }
    if mf_type_clean and mf_type_clean.upper() not in generic_material_types:
        return mf_type_clean
    return _text(mf_noun) or _text(part_description)


def _normalize_metadata(metadata_by_part: Mapping[str, Any] | None) -> dict[str, dict[str, Any]]:
    if not metadata_by_part:
        return {}
    normalized: dict[str, dict[str, Any]] = {}
    for raw_key, raw_value in metadata_by_part.items():
        key = _text(raw_key).upper()
        if not key:
            continue
        if not isinstance(raw_value, Mapping):
            raise ConversionError(f'Metadata for "{raw_key}" must be an object/record.')
        normalized[key] = dict(raw_value)
    return normalized


def load_metadata(path: str | Path) -> dict[str, dict[str, Any]]:
    """Load optional enrichment from JSON or CSV, keyed by part_no."""
    source = Path(path)
    if not source.is_file():
        raise ConversionError(f'Metadata file not found: "{source}".')
    try:
        if source.suffix.lower() == ".csv":
            with source.open("r", encoding="utf-8-sig", newline="") as handle:
                rows = list(csv.DictReader(handle))
            result: dict[str, dict[str, Any]] = {}
            for row_no, row in enumerate(rows, start=2):
                part_no = _text(row.get("part_no") or row.get("Part No.") or row.get("PART NUMBER"))
                if not part_no:
                    raise ConversionError(f"Metadata CSV row {row_no} has no part_no.")
                record = dict(row)
                if "beveled" in record:
                    value = _text(record["beveled"]).lower()
                    record["beveled"] = value in {"true", "yes", "1", "y"}
                result[part_no] = record
            return result
        if source.suffix.lower() == ".json":
            payload = json.loads(source.read_text(encoding="utf-8-sig"))
            if isinstance(payload, dict) and "metadata_by_part" in payload:
                payload = payload["metadata_by_part"]
            if isinstance(payload, list):
                result = {}
                for index, record in enumerate(payload, start=1):
                    if not isinstance(record, Mapping) or not _text(record.get("part_no")):
                        raise ConversionError(f"Metadata JSON record {index} has no part_no.")
                    result[_text(record["part_no"])] = dict(record)
                return result
            if isinstance(payload, Mapping):
                return {str(k): dict(v) for k, v in payload.items()}
            raise ConversionError("Metadata JSON must be an object keyed by part_no or a list of records.")
    except (OSError, csv.Error, json.JSONDecodeError, TypeError) as exc:
        raise ConversionError(f'Could not read metadata file "{source.name}": {exc}') from exc
    raise ConversionError("Metadata must be a .json or .csv file.")


def _extract_hull_number(path: Path, first_cell: Any) -> str:
    for candidate in [_text(first_cell), path.stem]:
        match = re.match(r"\s*(\d{3,})(?:\D|$)", candidate)
        if match:
            return match.group(1)
    return ""


def _find_bop_header(ws: Any) -> tuple[int, dict[str, int]]:
    for row_no, values in enumerate(
        ws.iter_rows(min_row=1, max_row=min(ws.max_row, 60), values_only=True), start=1
    ):
        normalized = [_header_key(value) for value in values]
        resolved: dict[str, int] = {}
        for canonical, aliases in BOP_HEADER_ALIASES.items():
            for index, value in enumerate(normalized, start=1):
                if value in aliases:
                    resolved[canonical] = index
                    break
        if {"bom_line", "part_no", "loa"}.issubset(resolved):
            return row_no, resolved
    raise ConversionError(
        f'The BOP sheet "{ws.title}" does not contain recognizable '
        "BOM Line, DYPN, and LOA reference headers."
    )


def _read_bops(
    paths: list[Path],
) -> tuple[dict[tuple[str, str], BOPEntry], list[Issue], set[str]]:
    index: dict[tuple[str, str], BOPEntry] = {}
    duplicate_counts: Counter[tuple[str, str]] = Counter()
    issues: list[Issue] = []
    bop_loas: set[str] = set()
    parsed_sheets = 0

    for path in paths:
        if not path.is_file():
            raise ConversionError(f'BOP workbook not found: "{path}".')
        if path.suffix.lower() not in {".xlsx", ".xlsm"}:
            raise ConversionError(f'"{path.name}" is not an .xlsx or .xlsm BOP workbook.')
        try:
            wb = load_workbook(path, data_only=True, read_only=True)
        except Exception as exc:
            raise ConversionError(f'Could not open BOP workbook "{path.name}": {exc}') from exc
        workbook_parsed = False
        for ws in wb.worksheets:
            try:
                header_row, columns = _find_bop_header(ws)
            except ConversionError:
                continue
            workbook_parsed = True
            parsed_sheets += 1
            first_cell = next(
                ws.iter_rows(min_row=1, max_row=1, max_col=1, values_only=True)
            )[0]
            hull_no = _extract_hull_number(path, first_cell)
            rows = ws.iter_rows(min_row=header_row + 1, values_only=True)
            for row_no, values in enumerate(rows, start=header_row + 1):
                part_no = _text(_row_value(values, columns, "part_no"))
                loa_raw = _text(_row_value(values, columns, "loa"))
                loa_match = re.search(r"(?i)LOA\d+", loa_raw)
                if not part_no or not loa_match:
                    continue
                loa = loa_match.group(0).upper()
                bom_line = _text(_row_value(values, columns, "bom_line"))
                mdlprt_col = columns.get("bom_line", 1) + 1
                mdlprt = _text(values[mdlprt_col - 1]) if mdlprt_col <= len(values) else ""
                if not mdlprt:
                    match = re.match(r"\s*([^/;]+)", bom_line)
                    mdlprt = _text(match.group(1)) if match else ""
                entry = BOPEntry(
                    source_file=path.name,
                    source_row=row_no,
                    hull_no=hull_no,
                    loa=loa,
                    part_no=part_no,
                    mdlprt=mdlprt,
                    mf_part_no=_text(_row_value(values, columns, "mf_part_no")),
                    quantity=_row_value(values, columns, "quantity"),
                    bom_line=bom_line,
                    revision=_text(_row_value(values, columns, "revision")),
                )
                key = (loa, part_no.upper())
                duplicate_counts[key] += 1
                bop_loas.add(loa)
                if key not in index:
                    index[key] = entry
                else:
                    existing = index[key]
                    comparable = (
                        existing.mdlprt,
                        existing.mf_part_no,
                        _text(existing.quantity),
                    )
                    incoming = (entry.mdlprt, entry.mf_part_no, _text(entry.quantity))
                    if comparable != incoming:
                        issues.append(
                            Issue(
                                "WARNING",
                                "CONFLICTING_BOP_ENTRY",
                                "Duplicate BOP rows disagree; the first row was retained.",
                                path.name,
                                row_no,
                                part_no,
                                f"{comparable} vs {incoming}",
                            )
                        )
        wb.close()
        if not workbook_parsed:
            raise ConversionError(
                f'"{path.name}" does not contain a recognizable BOP data sheet.'
            )

    if not parsed_sheets or not index:
        raise ConversionError("The selected BOP workbook(s) contain no scoped part rows.")
    for (loa, part_no), count in duplicate_counts.items():
        if count > 1:
            first = index[(loa, part_no)]
            issues.append(
                Issue(
                    "INFO",
                    "DUPLICATE_BOP_ENTRY",
                    f"BOP contains this LOA/part {count} times; it is included once.",
                    first.source_file,
                    first.source_row,
                    first.part_no,
                    loa,
                )
            )
    return index, issues, bop_loas


def _apply_metadata(record: PartRecord, metadata: Mapping[str, Any], issues: list[Issue]) -> None:
    supplied = metadata.get(record.part_no.upper())
    if supplied is None:
        missing = []
        if not record.mdlprt:
            missing.append("mdlprt")
        if record.category == "shape" and not record.length:
            missing.append("length")
        if missing:
            issues.append(
                Issue(
                    "INFO",
                    "MISSING_ENRICHMENT",
                    f"Optional enrichment is blank: {', '.join(missing)}.",
                    record.source_file,
                    record.source_row,
                    record.part_no,
                )
            )
        return

    record.mdlprt = _text(supplied.get("mdlprt")) or record.mdlprt
    if supplied.get("width") not in (None, ""):
        record.width = supplied["width"]
    if supplied.get("length") not in (None, ""):
        record.length = supplied["length"]
        record.length_inches = _length_to_inches(record.length)
    if _text(supplied.get("other_info")):
        record.other_info = _text(supplied["other_info"])
    if supplied.get("beveled") is True:
        record.other_info = "; ".join(x for x in [record.other_info, "BEVELED"] if x)


def _read_epl(
    path: Path,
    metadata: Mapping[str, Any],
    config: Mapping[str, Any],
    bop_index: Mapping[tuple[str, str], BOPEntry],
    seen_bop_keys: set[tuple[str, str]],
) -> EPLResult:
    if not path.is_file():
        raise ConversionError(f'Input workbook not found: "{path}".')
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ConversionError(f'"{path.name}" is not an .xlsx or .xlsm workbook.')
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        raise ConversionError(f'Could not open "{path.name}" as an Excel workbook: {exc}') from exc
    if "EPL" not in wb.sheetnames:
        raise ConversionError(f'"{path.name}" does not contain a sheet named EPL.')

    ws = wb["EPL"]
    header_row, columns = _find_header_row(ws)
    loa = _extract_loa(path)
    raw_records: list[dict[str, Any]] = []
    part_numbers: list[str] = []
    assemblies = 0
    source_rows = 0
    out_of_scope = 0

    rows = ws.iter_rows(min_row=header_row + 1, values_only=True)
    for row_no, values in enumerate(rows, start=header_row + 1):
        part_no = _text(_row_value(values, columns, "part_no"))
        description = _text(_row_value(values, columns, "part_description"))
        if not part_no and not description:
            continue
        source_rows += 1
        if description.upper() == "ASSEMBLY":
            assemblies += 1
            continue
        bop_key = (loa, part_no.upper())
        bop_entry = bop_index.get(bop_key)
        if bop_entry is None:
            out_of_scope += 1
            continue
        seen_bop_keys.add(bop_key)
        part_numbers.append(part_no)
        raw_records.append(
            {
                "row_no": row_no,
                "part_no": part_no,
                "part_description": description,
                "quantity": _row_value(values, columns, "quantity"),
                "mf_part_no": _text(_row_value(values, columns, "mf_part_no")),
                "mf_noun": _text(_row_value(values, columns, "mf_noun_name")),
                "mf_type": _text(_row_value(values, columns, "mf_type")),
                "material_specification": _text(
                    _row_value(values, columns, "material_specification")
                ),
                "material_description": _text(
                    _row_value(values, columns, "material_description")
                ),
                "thickness": _text(_row_value(values, columns, "thickness")),
                "width": _row_value(values, columns, "width") or "",
                "length": _row_value(values, columns, "length") or "",
                "unit_weight": _row_value(values, columns, "unit_weight"),
                "bop_entry": bop_entry,
            }
        )

    drawing_id = _drawing_identifier(part_numbers)
    result = EPLResult(
        source_path=path,
        loa=loa,
        drawing_id=drawing_id,
        assemblies_omitted=assemblies,
        source_rows=source_rows,
        bop_selected_rows=len(raw_records),
        out_of_scope_rows=out_of_scope,
    )
    duplicate_counts = Counter(x["part_no"] for x in raw_records if x["part_no"])
    for part_no, count in duplicate_counts.items():
        if count > 1:
            result.issues.append(
                Issue(
                    "WARNING",
                    "DUPLICATE_PART_NUMBER",
                    f"Part number occurs {count} times; every source row was retained.",
                    path.name,
                    None,
                    part_no,
                )
            )

    for raw in raw_records:
        bop_entry = raw["bop_entry"]
        category = _classify(
            raw["part_description"], raw["mf_noun"], raw["mf_type"], config
        )
        if not category:
            result.issues.append(
                Issue(
                    "WARNING",
                    "UNCLASSIFIED_ITEM",
                    "Non-assembly item was not exported because its stock category is not mapped.",
                    path.name,
                    raw["row_no"],
                    raw["part_no"],
                    " / ".join(
                        x
                        for x in [raw["part_description"], raw["mf_noun"], raw["mf_type"]]
                        if x
                    ),
                )
            )
            continue

        parsed = _parse_explicit_dimensions(raw["material_description"])
        material, mapped = _translate_material(
            raw["material_specification"],
            raw["material_description"],
            raw["mf_type"],
            config,
        )
        record = PartRecord(
            source_file=path.name,
            source_row=raw["row_no"],
            loa=loa,
            drawing_id=drawing_id,
            hull_no=bop_entry.hull_no,
            quantity=raw["quantity"],
            mf_part_no=raw["mf_part_no"],
            part_no=raw["part_no"],
            description=_display_description(
                raw["part_description"], raw["mf_noun"], raw["mf_type"]
            ),
            part_description=raw["part_description"],
            material_grade=material,
            material_specification=raw["material_specification"],
            material_description=raw["material_description"],
            thickness=raw["thickness"] or parsed.get("thickness", ""),
            width=raw["width"] or parsed.get("width", ""),
            length=raw["length"] or parsed.get("length", ""),
            size_thickness=parsed.get("size_thickness", "")
            or raw["thickness"]
            or parsed.get("thickness", ""),
            unit_weight=raw["unit_weight"] if isinstance(raw["unit_weight"], (int, float)) else None,
            mdlprt=bop_entry.mdlprt,
            category=category,
        )
        if (
            bop_entry.mf_part_no
            and record.mf_part_no
            and bop_entry.mf_part_no.upper() != record.mf_part_no.upper()
        ):
            result.issues.append(
                Issue(
                    "WARNING",
                    "BOP_EPL_MF_PART_MISMATCH",
                    "BOP and EPL MF part numbers differ; the EPL value was retained.",
                    path.name,
                    raw["row_no"],
                    raw["part_no"],
                    f"BOP={bop_entry.mf_part_no}; EPL={record.mf_part_no}",
                )
            )
        record.length_inches = _length_to_inches(record.length)
        if record.unit_weight is not None and record.length_inches is not None:
            qty = record.quantity if isinstance(record.quantity, (int, float)) else 1
            record.total_weight = round(record.unit_weight * (record.length_inches / 12) * qty, 3)

        if not mapped:
            result.issues.append(
                Issue(
                    "WARNING",
                    "UNMAPPED_MATERIAL",
                    "Material was preserved as supplied but has no translation rule.",
                    path.name,
                    raw["row_no"],
                    raw["part_no"],
                    " | ".join(
                        x
                        for x in [
                            raw["mf_type"],
                            raw["material_specification"],
                            raw["material_description"],
                        ]
                        if x
                    ),
                )
            )
        _apply_metadata(record, metadata, result.issues)
        (result.plates if category == "plate" else result.shapes).append(record)

    wb.close()
    return result


def _unique_sheet_name(base: str, used: set[str]) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]", "-", base).strip("'") or "EPL"
    cleaned = cleaned[:31]
    candidate = cleaned
    number = 2
    while candidate.casefold() in used:
        suffix = f" ({number})"
        candidate = cleaned[: 31 - len(suffix)] + suffix
        number += 1
    used.add(candidate.casefold())
    return candidate


def _style_sheet(ws: Any, headers: list[str], status_column: int) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(ws.max_row, 1)}"
    ws.row_dimensions[1].height = 25
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BODY_BORDER
    for row_no in range(2, ws.max_row + 1):
        fill = GRAY_FILL if row_no % 2 == 0 else WHITE_FILL
        ws.row_dimensions[row_no].height = 22
        for col_no in range(1, len(headers) + 1):
            cell = ws.cell(row_no, col_no)
            cell.fill = GREEN_FILL if col_no == status_column else fill
            cell.font = BODY_FONT
            cell.border = BODY_BORDER
            cell.alignment = Alignment(
                horizontal="left"
                if headers[col_no - 1] in {"Description", "other info"}
                else "center",
                vertical="center",
                wrap_text=headers[col_no - 1] == "other info",
            )
            if headers[col_no - 1] in {
                "LOA",
                "MF Part No.",
                "MDLPRT",
                "Part No.",
                "Drawing ID",
            }:
                cell.number_format = "@"

    widths = {
        "Item #": 10,
        "LOA": 15,
        "MF Part No.": 19,
        "Quantity": 12,
        "Description": 20,
        "MDLPRT": 21,
        "Part No.": 22,
        "Material Grade": 26,
        "THICKNESS": 16,
        "Size/Thickness": 18,
        "Width": 14,
        "Length": 15,
        "Length (in.)": 14,
        "Weight (lbs.)": 16,
        "Unit Weight (lbs/ft.)": 20,
        "Total Weight (lbs.)": 20,
        "STATUS": 13,
        "other info": 20,
        "Source File": 34,
        "Source Rows": 15,
        "Assemblies Omitted": 18,
        "Plates Exported": 17,
        "Shapes Exported": 17,
        "Issues": 12,
        "Severity": 12,
        "Code": 25,
        "Source Row": 12,
        "Part No.": 22,
        "Value": 55,
        "Message": 55,
    }
    for col_no, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_no)].width = widths.get(header, 15)
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = "landscape"
    ws.print_title_rows = "1:1"


def _plate_row(index: int, item: PartRecord) -> list[Any]:
    return [
        None,
        index,
        item.hull_no,
        item.loa,
        item.mf_part_no,
        item.quantity,
        item.description,
        item.part_no,
        item.material_grade,
        item.thickness,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        item.other_info,
    ]


def _shape_row(index: int, item: PartRecord) -> list[Any]:
    return [
        index,
        item.loa,
        item.mf_part_no,
        item.quantity,
        item.description,
        item.mdlprt,
        item.part_no,
        item.material_grade,
        item.size_thickness,
        item.length,
        item.length_inches,
        item.unit_weight,
        item.total_weight,
        item.status,
        item.other_info,
    ]


def _copy_cell_style(source: Any, target: Any) -> None:
    target._style = copy.copy(source._style)
    target.font = copy.copy(source.font)
    target.fill = copy.copy(source.fill)
    target.border = copy.copy(source.border)
    target.alignment = copy.copy(source.alignment)
    target.number_format = source.number_format
    target.protection = copy.copy(source.protection)


def _apply_default_plate_layout(ws: Any) -> None:
    blue = PatternFill("solid", fgColor="0070C0")
    white_font = Font(name="Arial Narrow", size=12, bold=True, color="FFFFFF")
    medium = Side(style="medium", color="000000")
    thin = Side(style="thin", color="000000")
    for col, value in enumerate(PLATE_TOP_HEADERS, start=2):
        ws.cell(2, col).value = value
    for col, value in enumerate(PLATE_PROCESS_HEADERS, start=2):
        ws.cell(3, col).value = value
    for col in list(range(2, 11)) + [19]:
        ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
    ws.merge_cells(start_row=2, start_column=11, end_row=2, end_column=18)
    for row in (2, 3):
        for col in range(2, 20):
            cell = ws.cell(row, col)
            cell.fill = blue
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=medium if col == 2 else thin,
                right=medium if col == 19 else thin,
                top=medium if row == 2 else thin,
                bottom=medium if row == 3 else thin,
            )
    widths = [3.85, 11.14, 11.14, 15.57, 19.57, 13, 20.71, 21.57, 21.14, 16]
    widths += [13.85, 14, 15, 14.14, 13.85, 14.28, 14.42, 14.57, 20]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    for row in range(1, 4):
        ws.row_dimensions[row].height = 15.75
    ws.sheet_view.zoomScale = 90


def _apply_plate_template_layout(ws: Any, template_ws: Any) -> None:
    for col in range(1, 20):
        source_dim = template_ws.column_dimensions[get_column_letter(col)]
        target_dim = ws.column_dimensions[get_column_letter(col)]
        target_dim.width = source_dim.width
        target_dim.hidden = source_dim.hidden
    for row in range(1, 4):
        ws.row_dimensions[row].height = template_ws.row_dimensions[row].height
        for col in range(1, 20):
            source = template_ws.cell(row, col)
            target = ws.cell(row, col, source.value)
            _copy_cell_style(source, target)
    for merged in template_ws.merged_cells.ranges:
        if merged.max_row <= 3 and merged.max_col <= 19:
            ws.merge_cells(str(merged))
    ws.sheet_view.showGridLines = template_ws.sheet_view.showGridLines
    ws.sheet_view.zoomScale = template_ws.sheet_view.zoomScale
    ws.page_margins = copy.copy(template_ws.page_margins)
    ws.page_setup = copy.copy(template_ws.page_setup)
    ws.sheet_properties.pageSetUpPr = copy.copy(template_ws.sheet_properties.pageSetUpPr)


def _style_plate_data_row(ws: Any, row_no: int, template_ws: Any | None) -> None:
    for col in range(1, 20):
        cell = ws.cell(row_no, col)
        if template_ws is not None:
            _copy_cell_style(template_ws.cell(4, col), cell)
            if 11 <= col <= 18 and cell.value in (None, ""):
                # A template example row may contain a green process "X".
                # Preserve its borders/alignment, not that row-specific process state.
                cell.fill = PatternFill(fill_type=None)
        else:
            cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style="medium", color="000000") if col == 2 else Side(style="thin", color="000000"),
                right=Side(style="medium", color="000000") if col == 19 else Side(style="thin", color="000000"),
                top=Side(style="thin", color="000000"),
                bottom=Side(style="thin", color="000000"),
            )
        if col in {4, 5, 8}:
            cell.number_format = "@"
    ws.row_dimensions[row_no].height = (
        template_ws.row_dimensions[4].height if template_ws is not None else 15.75
    )


def _build_plate_workbook(
    inputs: list[EPLResult], template_path: Path | None = None
) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    used: set[str] = set()
    template_wb = None
    template_ws = None
    if template_path is not None:
        try:
            template_wb = load_workbook(template_path, data_only=False, read_only=False)
            template_ws = next(
                (sheet for sheet in template_wb.worksheets if sheet.sheet_state == "visible"),
                template_wb.worksheets[0],
            )
        except Exception as exc:
            raise ConversionError(f'Could not read Plates template "{template_path.name}": {exc}') from exc
    for item in inputs:
        base = " - ".join(x for x in [item.loa, item.drawing_id] if x)
        ws = wb.create_sheet(_unique_sheet_name(base, used))
        if template_ws is not None:
            _apply_plate_template_layout(ws, template_ws)
        else:
            _apply_default_plate_layout(ws)
        for index, record in enumerate(item.plates, start=1):
            row_no = index + 3
            for col, value in enumerate(_plate_row(index, record), start=1):
                ws.cell(row_no, col).value = value
            _style_plate_data_row(ws, row_no, template_ws)
        last_row = max(ws.max_row, 3)
        ws.auto_filter.ref = f"B3:S{last_row}"
        ws.freeze_panes = "B4"
        ws.print_title_rows = "2:3"
    if template_wb is not None:
        template_wb.close()
    return wb


def _build_shape_workbook(inputs: list[EPLResult]) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    used: set[str] = set()
    for item in inputs:
        base = " - ".join(x for x in [item.loa, item.drawing_id] if x)
        ws = wb.create_sheet(_unique_sheet_name(base, used))
        ws.append([None] * (len(SHAPE_HEADERS) + 1))
        ws.append([None] + SHAPE_HEADERS)
        for index, record in enumerate(item.shapes, start=1):
            ws.append([None] + _shape_row(index, record))
        ws.delete_rows(1)
        ws.insert_rows(1)
        # Headers remain on row 2, matching the supplied Shapes workbook.
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "B3"
        ws.auto_filter.ref = f"B2:P{max(ws.max_row, 2)}"
        for col in range(2, len(SHAPE_HEADERS) + 2):
            cell = ws.cell(2, col)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BODY_BORDER
            ws.column_dimensions[get_column_letter(col)].width = 15
        for row in range(3, ws.max_row + 1):
            fill = GRAY_FILL if row % 2 == 1 else WHITE_FILL
            for col in range(2, len(SHAPE_HEADERS) + 2):
                cell = ws.cell(row, col)
                cell.fill = GREEN_FILL if col == 15 else fill
                cell.font = BODY_FONT
                cell.border = BODY_BORDER
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if col in {3, 4, 7, 8}:
                    cell.number_format = "@"
        widths = [11.14, 15.57, 19.57, 13, 16.85, 21, 28.14, 29.71, 16, 13.85, 13.85, 19, 20.57, 13.71, 14.42]
        for col, width in enumerate(widths, start=2):
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.page_setup.orientation = "landscape"
        ws.print_title_rows = "2:2"
    return wb


def _build_report_workbook(inputs: list[EPLResult], issues: list[Issue]) -> Workbook:
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    headers = [
        "Source File",
        "LOA",
        "Drawing ID",
        "Source Rows",
        "Assemblies Omitted",
        "BOP Selected",
        "EPL Out of Scope",
        "Plates Exported",
        "Shapes Exported",
        "Issues",
    ]
    summary.append(headers)
    for item in inputs:
        summary.append(
            [
                item.source_path.name,
                item.loa,
                item.drawing_id,
                item.source_rows,
                item.assemblies_omitted,
                item.bop_selected_rows,
                item.out_of_scope_rows,
                len(item.plates),
                len(item.shapes),
                len(item.issues),
            ]
        )
    _style_sheet(summary, headers, status_column=-1)

    issue_sheet = wb.create_sheet("Issues")
    issue_sheet.append(ISSUE_HEADERS)
    for issue in issues:
        issue_sheet.append(
            [
                issue.severity,
                issue.code,
                issue.source_file,
                issue.source_row,
                issue.part_no,
                issue.value,
                issue.message,
            ]
        )
    _style_sheet(issue_sheet, ISSUE_HEADERS, status_column=-1)
    issue_sheet.column_dimensions["F"].width = 65
    issue_sheet.column_dimensions["G"].width = 65
    for row in range(2, issue_sheet.max_row + 1):
        severity = issue_sheet.cell(row, 1).value
        issue_sheet.cell(row, 1).fill = (
            RED_FILL if severity == "ERROR" else AMBER_FILL if severity == "WARNING" else GRAY_FILL
        )
    return wb


def _atomic_save_workbook(workbook: Workbook, destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            prefix=f".{destination.stem}-", suffix=".xlsx", dir=destination.parent, delete=False
        ) as handle:
            temporary = Path(handle.name)
        workbook.save(temporary)
        os.replace(temporary, destination)
    except PermissionError as exc:
        raise ConversionError(
            f'Could not write "{destination.name}". Close it in Excel and try again.'
        ) from exc
    except OSError as exc:
        raise ConversionError(f'Could not write "{destination}": {exc}') from exc
    finally:
        workbook.close()
        if temporary and temporary.exists():
            temporary.unlink(missing_ok=True)


def convert_epls(
    epl_paths: Iterable[str | Path],
    output_dir: str | Path,
    output_base: str,
    metadata_by_part: Mapping[str, Any] | None = None,
    bop_paths: Iterable[str | Path] | None = None,
    plate_template_path: str | Path | None = None,
) -> ConversionResult:
    """Convert BOP-scoped EPL rows into new Plates/Shapes workbooks."""
    paths = [Path(path).expanduser().resolve() for path in epl_paths]
    if not paths:
        raise ConversionError("Select at least one EPL workbook.")
    bpaths = [Path(path).expanduser().resolve() for path in (bop_paths or [])]
    if not bpaths:
        raise ConversionError(
            "Select at least one BOP workbook. The BOP determines which EPL parts are in scope."
        )
    template_path = (
        Path(plate_template_path).expanduser().resolve()
        if plate_template_path
        else None
    )
    if template_path is not None and not template_path.is_file():
        raise ConversionError(f'Plates template not found: "{template_path}".')
    output = Path(output_dir).expanduser().resolve()
    base = _safe_output_base(output_base)
    metadata = _normalize_metadata(metadata_by_part)
    try:
        config = load_config()
    except ValueError as exc:
        raise ConversionError(str(exc)) from exc

    bop_index, bop_issues, bop_loas = _read_bops(bpaths)
    seen_bop_keys: set[tuple[str, str]] = set()
    inputs = [
        _read_epl(path, metadata, config, bop_index, seen_bop_keys) for path in paths
    ]
    all_issues = list(bop_issues) + [issue for item in inputs for issue in item.issues]
    selected_loas = {item.loa for item in inputs}
    for loa in sorted(bop_loas - selected_loas):
        count = sum(1 for key in bop_index if key[0] == loa)
        all_issues.append(
            Issue(
                "WARNING",
                "BOP_LOA_WITHOUT_EPL",
                f"BOP contains {count} scoped part(s) for this LOA, but no matching EPL was supplied.",
                ", ".join(path.name for path in bpaths),
                None,
                "",
                loa,
            )
        )
    for key, entry in bop_index.items():
        if key[0] in selected_loas and key not in seen_bop_keys:
            all_issues.append(
                Issue(
                    "WARNING",
                    "BOP_PART_NOT_IN_EPL",
                    "BOP-scoped part was not found in the supplied EPL for this LOA.",
                    entry.source_file,
                    entry.source_row,
                    entry.part_no,
                    entry.loa,
                )
            )

    plates_path = output / f"{base} - Plates.xlsx"
    shapes_path = output / f"{base} - Shapes.xlsx"
    report_path = output / f"{base} - Conversion Report.xlsx"
    report_json_path = output / f"{base} - Conversion Report.json"

    _atomic_save_workbook(_build_plate_workbook(inputs, template_path), plates_path)
    _atomic_save_workbook(_build_shape_workbook(inputs), shapes_path)
    _atomic_save_workbook(_build_report_workbook(inputs, all_issues), report_path)

    result = ConversionResult(
        plates_path=plates_path,
        shapes_path=shapes_path,
        report_path=report_path,
        report_json_path=report_json_path,
        inputs=inputs,
        issues=all_issues,
        bop_files=bpaths,
    )
    try:
        report_json_path.write_text(
            json.dumps(result.to_dict(), indent=2, ensure_ascii=False), encoding="utf-8"
        )
    except PermissionError as exc:
        raise ConversionError(
            f'Could not write "{report_json_path.name}". Close it and try again.'
        ) from exc
    return result
