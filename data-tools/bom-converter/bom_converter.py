"""
BOM to Parts List Converter

Double-click this file to open the graphical converter. The program lets a user
choose which source BOM column feeds each column in the Parts List template.
"""

import argparse
import json
import os
import re
import tkinter as tk
from copy import copy
from fractions import Fraction
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter, range_boundaries


APP_TITLE = "BOM to Parts List Converter"
IGNORE_MAPPING = "— Ignore this column —"
HEADER_SCAN_ROWS = 50
PREVIEW_ROWS = 100
SETTINGS_FILE = Path(__file__).resolve().with_name("bom_converter_mapping.json")

STANDARD_TEMPLATE_HEADERS = {
    "ORDER #",
    "PART NUMBER",
    "DESCRIPTION",
    "QUANTITY",
    "TOTAL QUANTITY",
    "THICKNESS SHAPE",
    "WIDTH",
    "LENGTH",
    "MATERIAL TYPE",
    "ASSEMBLY",
}

# Suggestions only. The user can override every one from the dropdowns.
SOURCE_SUGGESTIONS = {
    "THICKNESS SHAPE": "THICKNESS SHAPE",
    "THICKNESS/SHAPE": "THICKNESS SHAPE",
    "SHAPE": "THICKNESS SHAPE",
    "STOCK SIZE": "THICKNESS SHAPE",
    "ENG MAT ID": "PART NUMBER",
    "ENG MATID": "PART NUMBER",
    "ENGINEERING MATERIAL ID": "PART NUMBER",
    "PART NUMBER": "PART NUMBER",
    "PART NO": "PART NUMBER",
    "MATL MASTER DESCRIPTION": "DESCRIPTION",
    "MATERIAL MASTER DESCRIPTION": "DESCRIPTION",
    "MATERIAL DESCRIPTION": "DESCRIPTION",
    "DESCRIPTION": "DESCRIPTION",
    "REQD QTY": "QUANTITY",
    "REQUIRED QTY": "QUANTITY",
    "QTY": "QUANTITY",
    "QUANTITY": "QUANTITY",
    "SIZE 1": "WIDTH",
    "WIDTH": "WIDTH",
    "SIZE 2": "LENGTH",
    "LENGTH": "LENGTH",
    "MATERIAL TYPE": "MATERIAL TYPE",
    "MTL TYPE": "MATERIAL TYPE",
    "MANUFACTURING DWG": "ASSEMBLY",
    "ASSEMBLY": "ASSEMBLY",
}

# Steel plate weights from the supplied lb/sf table. The text values are kept
# exactly as they should appear in the Parts List THICKNESS/SHAPE column.
PLATE_WEIGHT_TO_THICKNESS = {
    7.65: "3/16",
    10.2: "1/4",
    12.75: "5/16",
    15.3: "3/8",
    17.85: "7/16",
    20.4: "1/2",
    22.95: "9/16",
    25.5: "5/8",
    28.05: "11/16",
    30.6: "3/4",
    33.15: "13/16",
    35.7: "7/8",
    40.8: "1",
    45.9: "1 1/8",
    51.0: "1 1/4",
    56.1: "1 3/8",
    61.2: "1 1/2",
    66.3: "1 5/8",
    71.4: "1 3/4",
    76.5: "1 7/8",
    81.6: "2",
    86.7: "2 1/8",
    91.8: "2 1/4",
    102.0: "2 1/2",
    112.2: "2 3/4",
    122.4: "3",
    132.6: "3 1/4",
    142.8: "3 1/2",
    153.0: "3 3/4",
    163.2: "4",
    173.4: "4 1/4",
    183.6: "4 1/2",
    204.0: "5",
    224.4: "5 1/2",
    244.8: "6",
    265.2: "6 1/2",
    285.6: "7",
    306.0: "7 1/2",
    326.4: "8",
    367.2: "9",
    408.0: "10",
}

# Longer stock-family names must be checked before the single-word fallback.
MATERIAL_DESCRIPTION_FAMILIES = (
    "RECTANGULAR TUBE",
    "SQUARE TUBE",
    "ROUND TUBE",
    "WIDE FLANGE",
    "ROUND BAR",
    "FLAT BAR",
    "I-BEAM",
    "I BEAM",
    "T-BAR",
    "T BAR",
)

MATERIAL_FAMILY_ALIASES = {
    "L": "ANGLE",
    "C": "CHANNEL",
    "W": "WIDE FLANGE",
    "WT": "TEE",
    "HSS": "SQUARE TUBE",
    "PL": "PLATE",
    "PLATE": "PLATE",
    "TEE": "TEE",
    "TBAR": "T-BAR",
    "T BAR": "T-BAR",
    "T-BAR": "T-BAR",
    "CHANNEL": "CHANNEL",
    "ANGLE": "ANGLE",
    "IBEAM": "I-BEAM",
    "I BEAM": "I-BEAM",
    "I-BEAM": "I-BEAM",
    "WIDE FLANGE": "WIDE FLANGE",
    "SQUARE TUBE": "SQUARE TUBE",
    "RECT TUBE": "RECTANGULAR TUBE",
    "RECTANGULAR TUBE": "RECTANGULAR TUBE",
    "ROUND TUBE": "ROUND TUBE",
    "FLAT BAR": "FLAT BAR",
    "ROUND BAR": "ROUND BAR",
    "PIPE": "PIPE",
}


class OutputRecord(dict):
    """A mapped row that can explicitly clear selected existing cells."""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.clear_destinations = set()


def normalize_header(value):
    """Normalize headings without changing the displayed text."""
    if value is None:
        return ""
    return re.sub(r"[^A-Z0-9]+", " ", str(value).upper()).strip()


def clean_excel_value(value):
    """Convert pandas missing values into blank Excel cells."""
    if pd.isna(value):
        return ""
    return value


def scan_header_rows(filepath, sheet_name=None, template_mode=False):
    """
    Detect the most likely header row.

    Source sheets favor rows containing many unique text cells. Template sheets
    strongly favor familiar Parts List headings.
    """
    workbook = openpyxl.load_workbook(
        filepath,
        read_only=True,
        data_only=True,
    )

    try:
        sheets = (
            [workbook[sheet_name]]
            if sheet_name
            else list(workbook.worksheets)
        )
        candidates = []

        for sheet in sheets:
            max_row = min(max(sheet.max_row, 1), HEADER_SCAN_ROWS)

            for row_number, row in enumerate(
                sheet.iter_rows(min_row=1, max_row=max_row),
                start=1,
            ):
                values = [
                    cell.value
                    for cell in row
                    if cell.value is not None and str(cell.value).strip()
                ]
                if not values:
                    continue

                normalized = [normalize_header(value) for value in values]
                unique_count = len(set(normalized))
                text_count = sum(isinstance(value, str) for value in values)
                known_count = sum(
                    value in STANDARD_TEMPLATE_HEADERS
                    for value in normalized
                )

                if template_mode:
                    score = known_count * 100 + text_count * 2 + unique_count
                else:
                    score = text_count * 3 + unique_count + known_count * 10

                candidates.append(
                    {
                        "sheet_name": sheet.title,
                        "header_row": row_number,
                        "score": score,
                        "known_count": known_count,
                        "nonempty_count": len(values),
                    }
                )

        if not candidates:
            raise ValueError("No nonempty rows were found in the workbook.")

        candidates.sort(
            key=lambda item: (
                item["score"],
                item["known_count"],
                item["nonempty_count"],
                -item["header_row"],
            ),
            reverse=True,
        )
        return candidates[0]
    finally:
        workbook.close()


def list_sheet_names(filepath):
    workbook = openpyxl.load_workbook(
        filepath,
        read_only=True,
        data_only=True,
    )
    try:
        return workbook.sheetnames
    finally:
        workbook.close()


def load_source_dataframe(filepath, sheet_name, header_row):
    dataframe = pd.read_excel(
        filepath,
        sheet_name=sheet_name,
        header=header_row - 1,
    )
    dataframe = dataframe.dropna(how="all")

    # Pandas makes duplicate headings unique with .1, .2, etc. Keeping those
    # names visible prevents an ambiguous mapping.
    dataframe.columns = [str(column).strip() for column in dataframe.columns]
    return dataframe


def read_template_layout(filepath):
    detected = scan_header_rows(filepath, template_mode=True)
    workbook = openpyxl.load_workbook(
        filepath,
        read_only=True,
        data_only=False,
    )

    try:
        sheet = workbook[detected["sheet_name"]]
        headers = []

        for cell in sheet[detected["header_row"]]:
            if cell.value is None or not str(cell.value).strip():
                continue
            headers.append(
                {
                    "name": str(cell.value).strip(),
                    "column": cell.column,
                    "letter": get_column_letter(cell.column),
                }
            )

        if not headers:
            raise ValueError("The detected template header row was empty.")
        normalized_names = [normalize_header(header["name"]) for header in headers]
        duplicates = sorted(
            name for name in set(normalized_names) if normalized_names.count(name) > 1
        )
        if duplicates:
            raise ValueError(
                "The template header row contains duplicate column names after "
                f"normalization: {', '.join(duplicates)}. Rename those columns "
                "so mappings cannot write to the wrong destination."
            )

        return {
            "sheet_name": detected["sheet_name"],
            "header_row": detected["header_row"],
            "data_start_row": detected["header_row"] + 1,
            "headers": headers,
        }
    finally:
        workbook.close()


def find_job_number(path):
    for folder in [path.parent, *path.parents]:
        match = re.search(r"(?<!\d)(\d{5})(?!\d)", folder.name)
        if match:
            return match.group(1)
    return None


def suggest_output_path(source_path):
    job_number = find_job_number(source_path)
    if job_number:
        filename = f"{job_number}_BULWARKS_MASTERS_PARTS_LIST.xlsx"
    else:
        filename = f"{source_path.stem}_PARTS_LIST.xlsx"
    return source_path.with_name(filename)


def parse_prefixes(text):
    return tuple(
        prefix.strip().upper()
        for prefix in re.split(r"[,; ]+", text)
        if prefix.strip()
    )


def find_template_header(template_headers, *normalized_names):
    """Return the template's displayed heading for any accepted name."""
    accepted = {normalize_header(name) for name in normalized_names}
    for header in template_headers or []:
        name = header["name"] if isinstance(header, dict) else str(header)
        if normalize_header(name) in accepted:
            return name
    return None


def clean_shape_text(value):
    """Make a stock designation compact without changing its meaning."""
    text = str(value or "").strip(" \t,;:-")
    text = text.replace("×", "X").replace("*", "X")
    text = re.sub(r"\s*[xX]\s*", "X", text)
    text = re.sub(r"\s+", " ", text)
    return text


def clean_structural_shape(value):
    """Normalize a complete structural designation for THICKNESS/SHAPE."""
    text = clean_shape_text(value)

    # PBOM dimensions are commonly padded to three decimals.  Keep meaningful
    # precision but render 5.000 as 5 so, for example, the complete tee stock
    # callout becomes 7.125X5X9.7# rather than being mistaken for plate weight.
    def trim_decimal(match):
        number = match.group(0)
        return number.rstrip("0").rstrip(".")

    return re.sub(r"(?<![\d.])\d+\.\d+(?![\d.])", trim_decimal, text)


def strip_trailing_material_text(value):
    """
    Remove common material-grade suffixes from a structural shape.

    The source description sometimes ends with values such as STL, ASTM-A36,
    AH36, A500, or 6061-T6. Those are material, not part of the stock shape.
    """
    text = clean_shape_text(value)
    material_start = re.search(
        r"(?:^|[\s,;])"
        r"(?=(?:STL|STEEL|ALUMINUM|ALUM|SS|ASTM\b|"
        r"AH[\s-]?\d|A[\s-]?\d{2,3}\b|"
        r"50(?:52|83)\b|60(?:61|63)\b|30[346]L?\b))",
        text,
        flags=re.IGNORECASE,
    )
    if material_start:
        text = text[:material_start.start()]
    return clean_shape_text(text)


def matching_plate_thickness(weight):
    """Return the table thickness for an lb/sf value, allowing tiny rounding."""
    try:
        numeric_weight = float(weight)
    except (TypeError, ValueError):
        return None

    for table_weight, thickness in PLATE_WEIGHT_TO_THICKNESS.items():
        if abs(numeric_weight - table_weight) <= 0.02:
            return thickness
    return None


def plate_thickness_from_text(value):
    """Convert a plate weight embedded in a description to plate thickness."""
    text = clean_shape_text(value)
    if not text:
        return ""

    # If the PBOM already supplies a fractional thickness, retain it.
    fraction = re.search(
        r"(?<![\d.])(?:(\d+)\s+)?(\d+)\s*/\s*(\d+)(?![\d.])",
        text,
    )
    if fraction:
        whole, numerator, denominator = fraction.groups()
        return (
            f"{whole} {numerator}/{denominator}"
            if whole
            else f"{numerator}/{denominator}"
        )

    # Prefer a number explicitly identified as plate weight.
    marked_weight = re.search(
        r"(?<![\d.])(\d+(?:\.\d+)?)\s*"
        r"(?:#|LB(?:S)?\s*/?\s*(?:SF|FT2|FT\^2))",
        text,
        flags=re.IGNORECASE,
    )
    if marked_weight:
        thickness = matching_plate_thickness(marked_weight.group(1))
        if thickness:
            return thickness

    # Some PBOM descriptions contain only "PLATE 10.2" with no weight unit.
    # In that case, use the first table-matching number after PLATE.
    for candidate in re.findall(r"(?<![\d.])\d+(?:\.\d+)?(?![\d.])", text):
        thickness = matching_plate_thickness(candidate)
        if thickness:
            return thickness

    # Keep the unrecognized value visible rather than silently discarding it.
    return strip_trailing_material_text(text)


def split_material_description(value):
    """
    Split a PBOM material description into DESCRIPTION and THICKNESS/SHAPE.

    Examples:
        PLATE 10.2#       -> PLATE | 1/4
        TEE 5X4.5#        -> TEE   | 5X4.5#
        CHANNEL 10X20 STL -> CHANNEL | 10X20
    """
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    if not text:
        return "", ""

    upper_text = text.upper()
    family = None
    remainder = ""

    for candidate in MATERIAL_DESCRIPTION_FAMILIES:
        if upper_text == candidate or upper_text.startswith(candidate + " "):
            family = candidate
            remainder = text[len(candidate):]
            break

    if family is None:
        match = re.match(r"([^\s,;:]+)(.*)", text)
        if not match:
            return text, ""
        family = match.group(1).strip(" \t,;:")
        remainder = match.group(2)

    description = family.upper()
    remainder = clean_shape_text(remainder)

    if normalize_header(description) in {"PLATE", "PL"}:
        description = "PLATE"
        shape = plate_thickness_from_text(remainder)
    else:
        shape = clean_structural_shape(strip_trailing_material_text(remainder))

    return description, shape


def material_code_from_token(value):
    """Return the short Parts List material code carried in a description."""
    normalized = normalize_header(value)
    if normalized in {"STL", "STEEL"}:
        return "STL"
    if normalized in {"AL", "ALUM", "ALUMINUM"}:
        return "AL"
    if normalized in {"SS", "STAINLESS", "STAINLESS STEEL"}:
        return "SS"
    return ""


def material_from_selected_value(value):
    """Keep the last comma-delimited material term, ignoring trailing Weight."""
    tokens = [token.strip() for token in str(value or "").split(",")]
    tokens = [token for token in tokens if token]
    while tokens and normalize_header(tokens[-1]) == "WEIGHT":
        tokens.pop()
    return tokens[-1] if tokens else ""


def parse_pbom_material_description(value):
    """
    Parse a comma-delimited PBOM raw-material description.

    The paired R rows in the supplied PBOM use:
        PLATE,STL,10.20# X120.000 X360.000
        TEE,STL,5.000D X2.690W X4.50#,STRL

    The first recognized stock family becomes DESCRIPTION, the material token
    becomes MATERIAL TYPE, and the next useful token becomes THICKNESS/SHAPE.
    """
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    if not text:
        return "", "", ""

    tokens = [
        token.strip()
        for token in re.split(r"[,;]", text)
        if token is not None and token.strip()
    ]

    family_index = None
    description = ""
    for index, token in enumerate(tokens):
        canonical = MATERIAL_FAMILY_ALIASES.get(normalize_header(token))
        if canonical:
            family_index = index
            description = canonical
            break

    if family_index is None:
        description, shape = split_material_description(text)
        return description, shape, ""

    material = ""
    shape_token = ""
    for token in tokens[family_index + 1:]:
        material_code = material_code_from_token(token)
        if material_code:
            if not material:
                material = material_code
            continue

        # STRL is a PBOM stock classification, not part of the shape.
        if normalize_header(token) in {"STRL", "STRUCTURAL"}:
            continue

        if not shape_token:
            shape_token = token

    if description == "PLATE":
        shape = plate_thickness_from_text(shape_token)
    else:
        shape = clean_structural_shape(shape_token)

    return description, shape, material


def parse_catia_name_spec(value):
    """Extract stock shape and grade from a CATIA ``NameSpecA`` value."""
    tokens = [token.strip() for token in str(value or "").split(",")]
    tokens = [token for token in tokens[1:] if token]
    if not tokens:
        return "", ""

    ignored_material_tokens = {"WEIGHT", "PL", "PLATE", "STRL", "STRUCTURAL"}
    material = material_from_selected_value(value)
    if (
        re.fullmatch(r"\d+(?:\.\d+)?", material)
        or re.match(
            r"^(?:BT|WT|W|C|L|HSS|PIPE|T)(?=\d)", material, re.IGNORECASE
        )
    ):
        material = ""
    shape = ""
    for index, token in enumerate(tokens):
        if normalize_header(token) in {"PL", "PLATE"}:
            if index + 1 < len(tokens):
                thickness = tokens[index + 1]
                try:
                    shape = str(Fraction(float(thickness)).limit_denominator(16))
                except (TypeError, ValueError, ZeroDivisionError):
                    shape = clean_structural_shape(thickness)
            break
        if re.match(
            r"^(?:BT|WT|W|C|L|HSS|PIPE|T)(?=\d)", token, re.IGNORECASE
        ):
            shape = clean_structural_shape(
                re.sub(r"^BT(?=\d)", "", token, flags=re.IGNORECASE)
            )
            break

    if material == shape or normalize_header(material) in ignored_material_tokens:
        material = ""
    return shape, material


def source_column_named(dataframe, *names):
    accepted = {normalize_header(name) for name in names}
    for column in dataframe.columns:
        if normalize_header(column) in accepted:
            return column
    return None


def mapped_destination_named(destination_sources, *names):
    accepted = {normalize_header(name) for name in names}
    for destination in destination_sources:
        if normalize_header(destination) in accepted:
            return destination
    return None


def nonblank_values(rows, source_column):
    if source_column is None:
        return []
    values = []
    for row in rows:
        value = clean_excel_value(row[source_column])
        if value == "":
            continue
        values.append(value)
    return values


def first_nonblank_value(rows, source_column):
    values = nonblank_values(rows, source_column)
    return values[0] if values else ""


def one_consistent_value(rows, source_column, part_number, field_name):
    """Return one repeated value or stop when one part has conflicting data."""
    values = nonblank_values(rows, source_column)
    if not values:
        return ""

    unique = {}
    for value in values:
        key = str(value).strip().upper()
        unique.setdefault(key, value)

    if len(unique) > 1:
        displayed = ", ".join(str(value) for value in unique.values())
        raise ValueError(
            f"{part_number} has conflicting {field_name} values in its "
            f"paired PBOM rows: {displayed}"
        )
    return next(iter(unique.values()))


def sum_quantity_values(rows, source_column, part_number):
    """Sum piece quantities from L rows only."""
    values = nonblank_values(rows, source_column)
    if not values:
        return ""

    total = 0.0
    for value in values:
        try:
            total += float(str(value).replace(",", "").strip())
        except ValueError as error:
            raise ValueError(
                f"{part_number} has a nonnumeric L-row quantity: {value}"
            ) from error

    return int(total) if total.is_integer() else total


def build_paired_lr_records(
    dataframe,
    destination_sources,
    prefixes,
    template_headers,
    split_descriptions,
):
    """Return one output record per part for hierarchical L/R PBOM exports."""
    category_source = source_column_named(dataframe, "ITEM CATEGORY")
    part_destination = mapped_destination_named(
        destination_sources,
        "PART NUMBER",
    )
    if category_source is None or part_destination is None:
        return None

    part_source = destination_sources[part_destination]
    groups = {}
    for _, row in dataframe.iterrows():
        part_number = str(clean_excel_value(row[part_source])).strip()
        if not part_number:
            continue
        if prefixes and not part_number.upper().startswith(prefixes):
            continue
        groups.setdefault(part_number, []).append(row)

    has_pair = any(
        {"L", "R"}.issubset(
            {
                str(clean_excel_value(row[category_source])).strip().upper()
                for row in rows
            }
        )
        for rows in groups.values()
    )
    if not has_pair:
        return None

    description_destination = mapped_destination_named(
        destination_sources,
        "DESCRIPTION",
    )
    quantity_destination = mapped_destination_named(
        destination_sources,
        "QUANTITY",
        "TOTAL QUANTITY",
        "QTY",
    )
    material_destination = mapped_destination_named(
        destination_sources,
        "MATERIAL TYPE",
        "MATERIAL",
    ) or find_template_header(template_headers, "MATERIAL TYPE", "MATERIAL")
    assembly_destination = mapped_destination_named(
        destination_sources,
        "ASSEMBLY",
    )
    width_destination = find_template_header(template_headers, "WIDTH")
    length_destination = find_template_header(template_headers, "LENGTH")
    shape_destination = find_template_header(
        template_headers,
        "THICKNESS SHAPE",
        "THICKNESS/SHAPE",
        "SHAPE",
    )

    description_source = (
        destination_sources.get(description_destination)
        if description_destination
        else source_column_named(
            dataframe,
            "MATL MASTER DESCRIPTION",
            "MATERIAL MASTER DESCRIPTION",
            "MATERIAL DESCRIPTION",
        )
    )
    quantity_source = source_column_named(
        dataframe,
        "REQD QTY",
        "REQUIRED QTY",
        "QUANTITY",
        "QTY",
    )
    size_1_source = source_column_named(dataframe, "SIZE 1", "WIDTH")
    size_2_source = source_column_named(dataframe, "SIZE 2")

    records = []
    for part_number, rows in groups.items():
        l_rows = [
            row
            for row in rows
            if str(clean_excel_value(row[category_source])).strip().upper()
            == "L"
        ]
        r_rows = [
            row
            for row in rows
            if str(clean_excel_value(row[category_source])).strip().upper()
            == "R"
        ]
        if not l_rows or not r_rows:
            raise ValueError(
                f"{part_number} is missing its paired "
                f"{'L' if not l_rows else 'R'} row."
            )

        record = OutputRecord()
        for destination, source in destination_sources.items():
            value = first_nonblank_value(l_rows, source)
            if value == "":
                value = first_nonblank_value(r_rows, source)
            record[destination] = value

        record[part_destination] = part_number

        raw_material_description = one_consistent_value(
            r_rows,
            description_source,
            part_number,
            "raw-material description",
        )
        description, derived_shape, derived_material = (
            parse_pbom_material_description(raw_material_description)
        )
        if material_destination is not None and material_destination in record:
            record[material_destination] = material_from_selected_value(
                record[material_destination]
            )
        if split_descriptions and description_destination is not None:
            record[description_destination] = description
            if shape_destination is not None:
                record[shape_destination] = derived_shape
            if material_destination is not None and derived_material:
                record[material_destination] = derived_material

        if quantity_destination is not None:
            record[quantity_destination] = sum_quantity_values(
                l_rows,
                quantity_source,
                part_number,
            )

        size_1 = one_consistent_value(
            r_rows,
            size_1_source,
            part_number,
            "Size 1",
        )
        size_2 = one_consistent_value(
            r_rows,
            size_2_source,
            part_number,
            "Size 2",
        )
        if description == "PLATE":
            if width_destination is not None:
                record[width_destination] = size_1
            if length_destination is not None:
                record[length_destination] = size_2
        else:
            if width_destination is not None:
                record[width_destination] = ""
                record.clear_destinations.add(width_destination)
            if length_destination is not None:
                record[length_destination] = size_1

        if assembly_destination is not None:
            record[assembly_destination] = one_consistent_value(
                l_rows,
                destination_sources[assembly_destination],
                part_number,
                "assembly",
            )

        records.append(record)

    return records


def sample_for_column(dataframe, column):
    values = []
    for value in dataframe[column]:
        cleaned = clean_excel_value(value)
        if cleaned == "":
            continue
        display = str(cleaned).replace("\n", " ").strip()
        if display and display not in values:
            values.append(display)
        if len(values) == 3:
            break

    if not values:
        return "(blank in sampled rows)"

    sample = "  |  ".join(values)
    return sample if len(sample) <= 75 else sample[:72] + "..."


def destination_by_normalized_name(template_headers):
    return {
        normalize_header(header["name"]): header["name"]
        for header in template_headers
    }


def suggest_mapping(source_column, template_headers):
    destinations = destination_by_normalized_name(template_headers)
    normalized_source = normalize_header(source_column)

    # Exact same heading is the safest suggestion.
    if normalized_source in destinations:
        return destinations[normalized_source]

    suggested_normalized = normalize_header(
        SOURCE_SUGGESTIONS.get(normalized_source, "")
    )
    return destinations.get(suggested_normalized, IGNORE_MAPPING)


def build_records(
    dataframe,
    mapping_rows,
    prefixes,
    template_headers=None,
    split_descriptions=True,
):
    """
    Build direct source-to-template records.

    mapping_rows contains one item per source column:
    {"source": <dataframe column>, "destination": <template heading>}.
    """
    destination_sources = {}

    for mapping in mapping_rows:
        destination = mapping["destination"]
        if destination == IGNORE_MAPPING:
            continue

        if destination in destination_sources:
            raise ValueError(
                f"Two source columns are mapped to '{destination}':\n"
                f"  • {destination_sources[destination]}\n"
                f"  • {mapping['source']}\n\n"
                "Choose only one source column for each destination."
            )
        destination_sources[destination] = mapping["source"]

    if not destination_sources:
        raise ValueError("Nothing is mapped. Select at least one destination.")

    part_destination = next(
        (
            destination
            for destination in destination_sources
            if normalize_header(destination) == "PART NUMBER"
        ),
        None,
    )

    if prefixes and part_destination is None:
        raise ValueError(
            "A part-number prefix filter is entered, but no source column is "
            "mapped to the template's PART NUMBER column."
        )

    description_destination = next(
        (
            destination
            for destination in destination_sources
            if normalize_header(destination) == "DESCRIPTION"
        ),
        None,
    )
    shape_destination = find_template_header(
        template_headers,
        "THICKNESS SHAPE",
        "THICKNESS/SHAPE",
        "SHAPE",
    )
    material_destination = find_template_header(
        template_headers, "MATERIAL TYPE", "MATERIAL"
    )
    catia_spec_source = source_column_named(
        dataframe, "NAMESPECA FROM CATIA", "NAMESPECA", "NAME SPEC A"
    )

    paired_records = build_paired_lr_records(
        dataframe,
        destination_sources,
        prefixes,
        template_headers,
        split_descriptions,
    )
    if paired_records is not None:
        if not paired_records:
            raise ValueError("No paired L/R PBOM records were found.")
        return paired_records

    records = []
    for _, source_row in dataframe.iterrows():
        record = {
            destination: clean_excel_value(source_row[source])
            for destination, source in destination_sources.items()
        }

        if not any(value != "" for value in record.values()):
            continue

        if prefixes:
            part_number = str(record.get(part_destination, "")).strip().upper()
            if not part_number.startswith(prefixes):
                continue

        if split_descriptions and description_destination is not None:
            raw_description = record.get(description_destination, "")
            if "," in str(raw_description or ""):
                description, derived_shape, _ = parse_pbom_material_description(
                    raw_description
                )
            else:
                description, derived_shape = split_material_description(
                    raw_description
                )
            record[description_destination] = description

            # An explicitly mapped, nonblank shape wins. Otherwise, populate
            # THICKNESS/SHAPE from the material description automatically.
            if (
                shape_destination is not None
                and (
                    shape_destination not in record
                    or record[shape_destination] == ""
                )
            ):
                record[shape_destination] = derived_shape

        # Material columns can contain an entire comma-delimited NameSpecA
        # value.  Drop a terminal "Weight" label and retain only the final
        # material term, regardless of which source column the user selected.
        if material_destination is not None and material_destination in record:
            record[material_destination] = material_from_selected_value(
                record[material_destination]
            )

        if catia_spec_source is not None:
            catia_shape, catia_material = parse_catia_name_spec(
                clean_excel_value(source_row[catia_spec_source])
            )
            if (
                shape_destination is not None
                and catia_shape
                and not record.get(shape_destination, "")
            ):
                record[shape_destination] = catia_shape
            if (
                material_destination is not None
                and catia_material
                and not record.get(material_destination, "")
            ):
                record[material_destination] = catia_material

        records.append(record)

    if not records:
        prefix_text = ", ".join(prefixes) if prefixes else "(none)"
        raise ValueError(
            "No source rows remained after mapping and filtering.\n"
            f"Current part-number prefix filter: {prefix_text}"
        )

    return records


def copy_row_style(sheet, source_row, destination_row):
    """Copy formatting only, never values or formulas."""
    if source_row == destination_row:
        return

    for column in range(1, sheet.max_column + 1):
        source_cell = sheet.cell(row=source_row, column=column)
        destination_cell = sheet.cell(row=destination_row, column=column)

        if source_cell.has_style:
            destination_cell._style = copy(source_cell._style)
        if source_cell.number_format:
            destination_cell.number_format = source_cell.number_format
        if source_cell.alignment:
            destination_cell.alignment = copy(source_cell.alignment)
        if source_cell.protection:
            destination_cell.protection = copy(source_cell.protection)

    if source_row in sheet.row_dimensions:
        source_dimension = sheet.row_dimensions[source_row]
        destination_dimension = sheet.row_dimensions[destination_row]
        destination_dimension.height = source_dimension.height
        destination_dimension.hidden = source_dimension.hidden


def extend_table_formatting(sheet, header_row, last_data_row):
    """Extend the template's Excel table through every generated record.

    Table banding is driven by the table reference, not by a fixed set of cell
    fills. Keeping the existing columns and extending only the final row lets
    engineers insert additional template columns without the converter relying
    on hard-coded positions or accidentally absorbing notes beside the table.
    """

    matching_tables = []
    for table in sheet.tables.values():
        min_column, min_row, max_column, max_row = range_boundaries(table.ref)
        if min_row == header_row:
            matching_tables.append(
                (table, min_column, min_row, max_column, max_row)
            )

    if not matching_tables:
        return None
    if len(matching_tables) > 1:
        names = ", ".join(item[0].name for item in matching_tables)
        raise ValueError(
            f"More than one Excel table starts on template header row "
            f"{header_row}: {names}. Keep one Parts List table on that row."
        )

    table, min_column, min_row, max_column, max_row = matching_tables[0]
    final_row = max(max_row, last_data_row)
    table.ref = (
        f"{get_column_letter(min_column)}{min_row}:"
        f"{get_column_letter(max_column)}{final_row}"
    )
    if table.autoFilter is not None:
        table.autoFilter.ref = table.ref
    return table.ref


def write_records_to_template(
    template_path,
    output_path,
    template_layout,
    records,
    update_existing=True,
):
    """Write mapped records into a new file or update an existing output."""
    use_existing_output = update_existing and output_path.exists()
    input_workbook_path = output_path if use_existing_output else template_path

    workbook = openpyxl.load_workbook(
        input_workbook_path,
        data_only=False,
    )

    try:
        sheet_name = template_layout["sheet_name"]
        sheet = (
            workbook[sheet_name]
            if sheet_name in workbook.sheetnames
            else workbook.active
        )
        header_row = template_layout["header_row"]
        data_start_row = template_layout["data_start_row"]

        output_columns = {}
        for cell in sheet[header_row]:
            normalized = normalize_header(cell.value)
            if normalized:
                if normalized in output_columns:
                    first = get_column_letter(output_columns[normalized])
                    second = get_column_letter(cell.column)
                    raise ValueError(
                        f"The output workbook has duplicate '{cell.value}' "
                        f"headers in columns {first} and {second}. Rename one "
                        "column before converting."
                    )
                output_columns[normalized] = cell.column

        mapped_columns = {}
        for destination in records[0]:
            normalized = normalize_header(destination)
            if normalized not in output_columns:
                raise ValueError(
                    f"The output workbook no longer contains the mapped "
                    f"column '{destination}'."
                )
            mapped_columns[destination] = output_columns[normalized]

        part_destination = next(
            (
                destination
                for destination in mapped_columns
                if normalize_header(destination) == "PART NUMBER"
            ),
            None,
        )
        part_column = (
            mapped_columns[part_destination]
            if part_destination is not None
            else None
        )

        existing_rows = {}
        if update_existing and part_column is not None:
            for row_number in range(data_start_row, sheet.max_row + 1):
                value = sheet.cell(row=row_number, column=part_column).value
                if value is None or not str(value).strip():
                    continue
                existing_rows.setdefault(str(value).strip(), row_number)

        columns_to_check = list(mapped_columns.values())
        next_row = data_start_row
        while any(
            sheet.cell(row=next_row, column=column).value not in (None, "")
            for column in columns_to_check
        ):
            next_row += 1

        style_source_row = data_start_row
        added = 0
        updated = 0
        last_written_row = data_start_row - 1

        for record in records:
            part_number = (
                str(record.get(part_destination, "")).strip()
                if part_destination is not None
                else ""
            )

            if part_number and part_number in existing_rows:
                row_number = existing_rows[part_number]
                updated += 1
            else:
                row_number = next_row
                next_row += 1
                copy_row_style(sheet, style_source_row, row_number)
                added += 1

                if part_number:
                    existing_rows[part_number] = row_number

            for destination, value in record.items():
                column = mapped_columns[destination]
                output_cell = sheet.cell(row=row_number, column=column)
                clear_destinations = getattr(
                    record,
                    "clear_destinations",
                    set(),
                )

                # A later routing/material row for the same part often leaves
                # fields blank. Do not let that erase a useful earlier value.
                if (
                    row_number in existing_rows.values()
                    and value == ""
                    and output_cell.value not in (None, "")
                    and destination not in clear_destinations
                ):
                    continue

                output_cell.value = value
            last_written_row = max(last_written_row, row_number)

        extend_table_formatting(sheet, header_row, last_written_row)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        temporary_output = output_path.with_name(
            f"{output_path.stem}.__writing__{output_path.suffix}"
        )

        try:
            workbook.save(temporary_output)
            os.replace(temporary_output, output_path)
        finally:
            if temporary_output.exists():
                temporary_output.unlink()

        return added, updated
    finally:
        workbook.close()


class MappingRow:
    def __init__(
        self,
        parent,
        row_number,
        source_column,
        sample,
        destinations,
        initial_destination,
    ):
        self.source_column = source_column
        self.destination = tk.StringVar(value=initial_destination)

        ttk.Label(
            parent,
            text=source_column,
            width=29,
            anchor="w",
        ).grid(row=row_number, column=0, sticky="ew", padx=(4, 8), pady=2)

        ttk.Label(
            parent,
            text=sample,
            width=55,
            anchor="w",
        ).grid(row=row_number, column=1, sticky="ew", padx=4, pady=2)

        combo = ttk.Combobox(
            parent,
            textvariable=self.destination,
            values=destinations,
            state="readonly",
            width=29,
        )
        combo.grid(row=row_number, column=2, sticky="ew", padx=(8, 4), pady=2)


class BomConverterApp:
    def __init__(self, root, close_when_done=False):
        self.root = root
        self.close_when_done = close_when_done
        self.root.title(APP_TITLE)
        self.root.geometry("1180x780")
        self.root.minsize(960, 650)

        self.source_path = tk.StringVar()
        self.template_path = tk.StringVar()
        self.output_path = tk.StringVar()
        self.sheet_name = tk.StringVar()
        self.source_header_row = tk.IntVar(value=1)
        self.prefixes = tk.StringVar(value="DS,DV")
        self.split_descriptions = tk.BooleanVar(value=True)
        self.update_existing = tk.BooleanVar(value=True)
        self.open_when_done = tk.BooleanVar(value=True)
        self.status = tk.StringVar(
            value="Choose a source BOM and a Parts List template to begin."
        )

        self.source_dataframe = None
        self.template_layout = None
        self.mapping_rows = []
        self.saved_mapping = self.load_saved_mapping()

        self.build_interface()

    def load_paths(self, source_path=None, output_path=None, template_path=None):
        """Preload Assistant-selected paths while retaining the setup GUI."""
        if source_path:
            self.source_path.set(str(source_path))
            sheets = list_sheet_names(Path(source_path))
            self.sheet_combo["values"] = sheets
            self.sheet_name.set(sheets[0])
            detected = scan_header_rows(
                Path(source_path), sheet_name=sheets[0], template_mode=False
            )
            self.source_header_row.set(detected["header_row"])
        if output_path:
            self.output_path.set(str(output_path))
        if template_path:
            self.template_path.set(str(template_path))
            self.load_template()
        if source_path:
            self.reload_source_columns()

    def build_interface(self):
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(1, weight=1)

        setup = ttk.LabelFrame(
            self.root,
            text="1. Choose the files",
            padding=10,
        )
        setup.grid(row=0, column=0, sticky="ew", padx=12, pady=(12, 6))
        setup.columnconfigure(1, weight=1)

        ttk.Label(setup, text="Source BOM:").grid(
            row=0, column=0, sticky="w", padx=(0, 8), pady=4
        )
        ttk.Entry(setup, textvariable=self.source_path).grid(
            row=0, column=1, columnspan=4, sticky="ew", pady=4
        )
        ttk.Button(setup, text="Browse…", command=self.choose_source).grid(
            row=0, column=5, padx=(8, 0), pady=4
        )

        ttk.Label(setup, text="Worksheet:").grid(
            row=1, column=0, sticky="w", padx=(0, 8), pady=4
        )
        self.sheet_combo = ttk.Combobox(
            setup,
            textvariable=self.sheet_name,
            state="readonly",
            width=28,
        )
        self.sheet_combo.grid(row=1, column=1, sticky="w", pady=4)
        self.sheet_combo.bind("<<ComboboxSelected>>", self.sheet_changed)

        ttk.Label(setup, text="Source header row:").grid(
            row=1, column=2, sticky="e", padx=(18, 8), pady=4
        )
        ttk.Spinbox(
            setup,
            from_=1,
            to=500,
            textvariable=self.source_header_row,
            width=7,
        ).grid(row=1, column=3, sticky="w", pady=4)
        ttk.Button(
            setup,
            text="Reload columns",
            command=self.reload_source_columns,
        ).grid(row=1, column=4, columnspan=2, sticky="e", pady=4)

        ttk.Label(setup, text="Parts List template:").grid(
            row=2, column=0, sticky="w", padx=(0, 8), pady=4
        )
        ttk.Entry(setup, textvariable=self.template_path).grid(
            row=2, column=1, columnspan=4, sticky="ew", pady=4
        )
        ttk.Button(setup, text="Browse…", command=self.choose_template).grid(
            row=2, column=5, padx=(8, 0), pady=4
        )

        ttk.Label(setup, text="Output file:").grid(
            row=3, column=0, sticky="w", padx=(0, 8), pady=4
        )
        ttk.Entry(setup, textvariable=self.output_path).grid(
            row=3, column=1, columnspan=4, sticky="ew", pady=4
        )
        ttk.Button(setup, text="Save as…", command=self.choose_output).grid(
            row=3, column=5, padx=(8, 0), pady=4
        )

        ttk.Label(setup, text="Part prefixes to keep:").grid(
            row=4, column=0, sticky="w", padx=(0, 8), pady=4
        )
        ttk.Entry(setup, textvariable=self.prefixes, width=22).grid(
            row=4, column=1, sticky="w", pady=4
        )
        ttk.Label(
            setup,
            text="Separate with commas. Leave blank to keep every row.",
        ).grid(row=4, column=2, columnspan=2, sticky="w", padx=(12, 0))
        ttk.Checkbutton(
            setup,
            text="Update matching part numbers in an existing output",
            variable=self.update_existing,
        ).grid(row=4, column=4, columnspan=2, sticky="e")

        ttk.Checkbutton(
            setup,
            text="Split material description into Description + Thickness/Shape",
            variable=self.split_descriptions,
        ).grid(row=5, column=0, columnspan=3, sticky="w", pady=(2, 0))
        ttk.Label(
            setup,
            text="Plate lb/sf weights use the supplied thickness table.",
        ).grid(row=5, column=3, columnspan=3, sticky="e", pady=(2, 0))

        mapping_frame = ttk.LabelFrame(
            self.root,
            text="2. Tell the program where each source column goes",
            padding=8,
        )
        mapping_frame.grid(
            row=1,
            column=0,
            sticky="nsew",
            padx=12,
            pady=6,
        )
        mapping_frame.columnconfigure(0, weight=1)
        mapping_frame.rowconfigure(1, weight=1)

        header = ttk.Frame(mapping_frame)
        header.grid(row=0, column=0, sticky="ew", padx=(0, 16))
        header.columnconfigure(0, weight=2)
        header.columnconfigure(1, weight=3)
        header.columnconfigure(2, weight=2)
        ttk.Label(header, text="Source column").grid(
            row=0, column=0, sticky="w", padx=4
        )
        ttk.Label(header, text="Example source values").grid(
            row=0, column=1, sticky="w", padx=4
        )
        ttk.Label(header, text="Send this column to…").grid(
            row=0, column=2, sticky="w", padx=4
        )

        canvas_holder = ttk.Frame(mapping_frame)
        canvas_holder.grid(row=1, column=0, sticky="nsew")
        canvas_holder.columnconfigure(0, weight=1)
        canvas_holder.rowconfigure(0, weight=1)

        self.mapping_canvas = tk.Canvas(
            canvas_holder,
            highlightthickness=0,
        )
        self.mapping_canvas.grid(row=0, column=0, sticky="nsew")
        scrollbar = ttk.Scrollbar(
            canvas_holder,
            orient="vertical",
            command=self.mapping_canvas.yview,
        )
        scrollbar.grid(row=0, column=1, sticky="ns")
        self.mapping_canvas.configure(yscrollcommand=scrollbar.set)

        self.mapping_inner = ttk.Frame(self.mapping_canvas)
        self.mapping_window = self.mapping_canvas.create_window(
            (0, 0),
            window=self.mapping_inner,
            anchor="nw",
        )
        self.mapping_inner.bind("<Configure>", self.mapping_frame_changed)
        self.mapping_canvas.bind("<Configure>", self.mapping_canvas_changed)
        self.mapping_canvas.bind_all("<MouseWheel>", self.mapping_mousewheel)

        self.empty_mapping_label = ttk.Label(
            self.mapping_inner,
            text=(
                "The source columns and dropdowns will appear here after both "
                "files are selected."
            ),
            padding=20,
        )
        self.empty_mapping_label.grid(row=0, column=0, columnspan=3)

        actions = ttk.Frame(self.root, padding=(12, 6, 12, 12))
        actions.grid(row=2, column=0, sticky="ew")
        actions.columnconfigure(0, weight=1)

        ttk.Label(actions, textvariable=self.status).grid(
            row=0, column=0, sticky="w"
        )
        ttk.Button(actions, text="Auto-map", command=self.auto_map).grid(
            row=0, column=1, padx=4
        )
        ttk.Button(actions, text="Clear mappings", command=self.clear_mappings).grid(
            row=0, column=2, padx=4
        )
        ttk.Button(actions, text="Preview output", command=self.preview_output).grid(
            row=0, column=3, padx=4
        )
        ttk.Checkbutton(
            actions,
            text="Open when finished",
            variable=self.open_when_done,
        ).grid(row=0, column=4, padx=(12, 8))
        ttk.Button(
            actions,
            text="Convert BOM",
            command=self.convert,
            style="Accent.TButton",
        ).grid(row=0, column=5, padx=(4, 0))

    def mapping_frame_changed(self, _event):
        self.mapping_canvas.configure(
            scrollregion=self.mapping_canvas.bbox("all")
        )

    def mapping_canvas_changed(self, event):
        self.mapping_canvas.itemconfigure(
            self.mapping_window,
            width=event.width,
        )

    def mapping_mousewheel(self, event):
        if self.mapping_canvas.winfo_exists():
            self.mapping_canvas.yview_scroll(
                int(-1 * (event.delta / 120)),
                "units",
            )

    def choose_source(self):
        selected = filedialog.askopenfilename(
            title="Choose the source BOM",
            filetypes=[
                ("Excel workbooks", "*.xlsx *.xlsm"),
                ("All files", "*.*"),
            ],
        )
        if not selected:
            return

        self.source_path.set(selected)

        try:
            sheets = list_sheet_names(Path(selected))
            self.sheet_combo["values"] = sheets
            self.sheet_name.set(sheets[0])

            detected = scan_header_rows(
                Path(selected),
                sheet_name=sheets[0],
                template_mode=False,
            )
            self.source_header_row.set(detected["header_row"])

            source_path = Path(selected)
            nearby_template = source_path.with_name("TEMP.xlsx")
            if not self.template_path.get() and nearby_template.exists():
                self.template_path.set(str(nearby_template))
                self.load_template()

            self.output_path.set(str(suggest_output_path(source_path)))
            self.reload_source_columns()
        except Exception as error:
            self.show_error("Could not load the source BOM", error)

    def sheet_changed(self, _event=None):
        try:
            detected = scan_header_rows(
                Path(self.source_path.get()),
                sheet_name=self.sheet_name.get(),
                template_mode=False,
            )
            self.source_header_row.set(detected["header_row"])
            self.reload_source_columns()
        except Exception as error:
            self.show_error("Could not load that worksheet", error)

    def reload_source_columns(self):
        if not self.source_path.get() or not self.sheet_name.get():
            return

        try:
            self.source_dataframe = load_source_dataframe(
                Path(self.source_path.get()),
                self.sheet_name.get(),
                self.source_header_row.get(),
            )
            self.status.set(
                f"Loaded {len(self.source_dataframe):,} source rows and "
                f"{len(self.source_dataframe.columns):,} source columns."
            )

            if self.template_path.get() and self.template_layout is None:
                self.load_template()
            self.rebuild_mapping_rows()
        except Exception as error:
            self.show_error("Could not read the source columns", error)

    def choose_template(self):
        selected = filedialog.askopenfilename(
            title="Choose the blank Parts List template",
            filetypes=[
                ("Excel workbooks", "*.xlsx *.xlsm"),
                ("All files", "*.*"),
            ],
        )
        if not selected:
            return

        self.template_path.set(selected)
        self.load_template()

    def load_template(self):
        try:
            self.template_layout = read_template_layout(
                Path(self.template_path.get())
            )
            self.status.set(
                f"Template headers found on "
                f"'{self.template_layout['sheet_name']}' row "
                f"{self.template_layout['header_row']}."
            )
            self.rebuild_mapping_rows()
        except Exception as error:
            self.template_layout = None
            self.show_error("Could not read the Parts List template", error)

    def choose_output(self):
        initial = self.output_path.get()
        selected = filedialog.asksaveasfilename(
            title="Choose the output Parts List",
            defaultextension=".xlsx",
            initialdir=str(Path(initial).parent) if initial else None,
            initialfile=Path(initial).name if initial else None,
            filetypes=[("Excel workbook", "*.xlsx")],
        )
        if selected:
            self.output_path.set(selected)

    def saved_destination_for(self, source_column, destinations):
        normalized_source = normalize_header(source_column)
        saved = self.saved_mapping.get("mapping", {}).get(normalized_source)
        if not saved:
            return None

        destination_lookup = {
            normalize_header(destination): destination
            for destination in destinations
        }
        return destination_lookup.get(saved)

    def rebuild_mapping_rows(self):
        if self.source_dataframe is None or self.template_layout is None:
            return

        for child in self.mapping_inner.winfo_children():
            child.destroy()
        self.mapping_rows = []

        template_headers = self.template_layout["headers"]
        destinations = [IGNORE_MAPPING] + [
            header["name"] for header in template_headers
        ]

        for row_number, source_column in enumerate(
            self.source_dataframe.columns
        ):
            saved = self.saved_destination_for(source_column, destinations)
            initial = saved or suggest_mapping(source_column, template_headers)
            mapping_row = MappingRow(
                self.mapping_inner,
                row_number,
                source_column,
                sample_for_column(self.source_dataframe, source_column),
                destinations,
                initial,
            )
            self.mapping_rows.append(mapping_row)

        self.mapping_inner.columnconfigure(0, weight=2)
        self.mapping_inner.columnconfigure(1, weight=3)
        self.mapping_inner.columnconfigure(2, weight=2)
        self.mapping_canvas.yview_moveto(0)

        mapped_count = sum(
            row.destination.get() != IGNORE_MAPPING
            for row in self.mapping_rows
        )
        self.status.set(
            f"Ready: {mapped_count} automatic mappings suggested. "
            "Review the dropdowns before converting."
        )

    def auto_map(self):
        if not self.mapping_rows or self.template_layout is None:
            messagebox.showinfo(
                APP_TITLE,
                "Choose the source BOM and template first.",
            )
            return

        for mapping_row in self.mapping_rows:
            mapping_row.destination.set(
                suggest_mapping(
                    mapping_row.source_column,
                    self.template_layout["headers"],
                )
            )
        self.status.set("Automatic mapping suggestions restored.")

    def clear_mappings(self):
        for mapping_row in self.mapping_rows:
            mapping_row.destination.set(IGNORE_MAPPING)
        self.status.set("All mappings cleared.")

    def current_mapping(self):
        return [
            {
                "source": row.source_column,
                "destination": row.destination.get(),
            }
            for row in self.mapping_rows
        ]

    def current_records(self):
        if self.source_dataframe is None:
            raise ValueError("Choose and load a source BOM first.")
        if self.template_layout is None:
            raise ValueError("Choose a Parts List template first.")

        return build_records(
            self.source_dataframe,
            self.current_mapping(),
            parse_prefixes(self.prefixes.get()),
            template_headers=self.template_layout["headers"],
            split_descriptions=self.split_descriptions.get(),
        )

    def preview_output(self):
        try:
            records = self.current_records()
        except Exception as error:
            self.show_error("Cannot build the preview", error)
            return

        preview = tk.Toplevel(self.root)
        preview.title(f"Mapped Output Preview — {len(records):,} rows")
        preview.geometry("1100x520")
        preview.rowconfigure(0, weight=1)
        preview.columnconfigure(0, weight=1)

        columns = list(records[0])
        tree = ttk.Treeview(
            preview,
            columns=columns,
            show="headings",
        )
        vertical = ttk.Scrollbar(
            preview,
            orient="vertical",
            command=tree.yview,
        )
        horizontal = ttk.Scrollbar(
            preview,
            orient="horizontal",
            command=tree.xview,
        )
        tree.configure(
            yscrollcommand=vertical.set,
            xscrollcommand=horizontal.set,
        )

        tree.grid(row=0, column=0, sticky="nsew")
        vertical.grid(row=0, column=1, sticky="ns")
        horizontal.grid(row=1, column=0, sticky="ew")

        for column in columns:
            tree.heading(column, text=column)
            tree.column(column, width=170, minwidth=100)

        for record in records[:PREVIEW_ROWS]:
            tree.insert(
                "",
                "end",
                values=[record[column] for column in columns],
            )

        footer = ttk.Label(
            preview,
            text=(
                f"Showing {min(len(records), PREVIEW_ROWS):,} of "
                f"{len(records):,} mapped rows. This preview does not change "
                "any files."
            ),
            padding=8,
        )
        footer.grid(row=2, column=0, columnspan=2, sticky="w")

    def save_mapping(self):
        mapping = {
            normalize_header(row.source_column): normalize_header(
                row.destination.get()
            )
            for row in self.mapping_rows
            if row.destination.get() != IGNORE_MAPPING
        }
        payload = {
            "mapping": mapping,
            "prefixes": self.prefixes.get(),
            "split_descriptions": self.split_descriptions.get(),
        }

        try:
            SETTINGS_FILE.write_text(
                json.dumps(payload, indent=2),
                encoding="utf-8",
            )
            self.saved_mapping = payload
        except OSError:
            # Conversion should still succeed if the network folder is read-only.
            pass

    def load_saved_mapping(self):
        try:
            if SETTINGS_FILE.exists():
                payload = json.loads(
                    SETTINGS_FILE.read_text(encoding="utf-8")
                )
                if isinstance(payload, dict):
                    prefixes = payload.get("prefixes")
                    if isinstance(prefixes, str):
                        self.prefixes.set(prefixes)
                    split_descriptions = payload.get("split_descriptions")
                    if isinstance(split_descriptions, bool):
                        self.split_descriptions.set(split_descriptions)
                    return payload
        except (OSError, json.JSONDecodeError):
            pass
        return {}

    def convert(self):
        try:
            records = self.current_records()

            template_path = Path(self.template_path.get()).resolve()
            output_path = Path(self.output_path.get()).resolve()
            source_path = Path(self.source_path.get()).resolve()

            if not self.output_path.get().strip():
                raise ValueError("Choose an output filename.")
            if output_path == template_path:
                raise ValueError(
                    "The output cannot overwrite the blank template. "
                    "Choose a different output filename."
                )
            if output_path == source_path:
                raise ValueError(
                    "The output cannot overwrite the source BOM. "
                    "Choose a different output filename."
                )

            if output_path.exists() and not self.update_existing.get():
                overwrite = messagebox.askyesno(
                    APP_TITLE,
                    f"{output_path.name} already exists.\n\n"
                    "Replace it with a new conversion from the template?",
                )
                if not overwrite:
                    return

            self.status.set(f"Converting {len(records):,} rows…")
            self.root.update_idletasks()

            added, updated = write_records_to_template(
                template_path,
                output_path,
                self.template_layout,
                records,
                update_existing=self.update_existing.get(),
            )
            self.save_mapping()

            self.status.set(
                f"Finished: {added:,} rows added and "
                f"{updated:,} matching parts updated."
            )
            messagebox.showinfo(
                APP_TITLE,
                f"Conversion complete.\n\n"
                f"Rows added: {added:,}\n"
                f"Matching parts updated: {updated:,}\n\n"
                f"Saved as:\n{output_path}",
            )

            if self.open_when_done.get() and os.name == "nt":
                os.startfile(output_path)
            if self.close_when_done:
                self.root.destroy()
        except PermissionError:
            messagebox.showerror(
                APP_TITLE,
                "The output workbook could not be saved. Close it in Excel "
                "and try again.",
            )
            self.status.set("Stopped: close the output workbook and retry.")
        except Exception as error:
            self.show_error("Conversion stopped", error)

    def show_error(self, title, error):
        messagebox.showerror(APP_TITLE, f"{title}.\n\n{error}")
        self.status.set(f"{title}: {error}")


def convert_workbook(source_path, output_path, template_path):
    """Run the same conversion engine non-interactively for trusted defaults.

    This entry point is used by the Job Assistant. It intentionally applies
    only the converter's existing safe mapping suggestions; ambiguous columns
    remain unmapped rather than being guessed.
    """
    source_path = Path(source_path).resolve()
    output_path = Path(output_path).resolve()
    template_path = Path(template_path).resolve()
    if output_path in {source_path, template_path}:
        raise ValueError("The output must not overwrite the source or template.")
    source = scan_header_rows(source_path, template_mode=False)
    dataframe = load_source_dataframe(
        source_path, source["sheet_name"], source["header_row"]
    )
    layout = read_template_layout(template_path)
    mappings = [
        {
            "source": column,
            "destination": suggest_mapping(column, layout["headers"]),
        }
        for column in dataframe.columns
    ]
    records = build_records(
        dataframe,
        mappings,
        parse_prefixes("DS,DV"),
        template_headers=layout["headers"],
        split_descriptions=True,
    )
    return write_records_to_template(
        template_path, output_path, layout, records, update_existing=True
    )


def main(argv=None):
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("source", nargs="?", type=Path)
    parser.add_argument("output", nargs="?", type=Path)
    parser.add_argument("template", nargs="?", type=Path)
    parser.add_argument("--gui", action="store_true")
    args = parser.parse_args(argv)
    supplied = [args.source, args.output, args.template]
    if any(supplied) and not args.gui:
        if not all(supplied):
            parser.error("source, output, and template must be supplied together")
        added, updated = convert_workbook(
            args.source, args.output, args.template
        )
        print(
            f"Conversion complete: {added} row(s) added and "
            f"{updated} matching part(s) updated. Output: {args.output}"
        )
        return 0

    root = tk.Tk()
    app = BomConverterApp(root, close_when_done=all(supplied))
    if all(supplied):
        try:
            app.load_paths(args.source, args.output, args.template)
        except Exception as error:
            app.show_error("Could not preload the Assistant-selected files", error)

    # Keep a reference for the lifetime of the Tk window.
    root.bom_converter_app = app
    root.mainloop()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
