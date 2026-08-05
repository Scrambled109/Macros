from __future__ import annotations

import importlib.util
import sys
import tempfile
import unittest
from pathlib import Path

import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo


MODULE_PATH = Path(__file__).resolve().parents[1] / "bom_converter.py"
SPEC = importlib.util.spec_from_file_location("bom_converter_template_tests", MODULE_PATH)
converter = importlib.util.module_from_spec(SPEC)
assert SPEC.loader is not None
sys.modules[SPEC.name] = converter
SPEC.loader.exec_module(converter)


class DynamicTemplateTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary.name)

    def tearDown(self) -> None:
        self.temporary.cleanup()

    def create_template(self, headers: list[str], table_end_row: int = 7) -> Path:
        path = self.root / "template.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Parts List"
        sheet["A1"] = "PARTS LIST"
        for column, header in enumerate(headers, start=1):
            sheet.cell(row=5, column=column, value=header)
        table = Table(
            displayName="PartsListTable",
            ref=(
                f"A5:{openpyxl.utils.get_column_letter(len(headers))}"
                f"{table_end_row}"
            ),
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium5",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
        workbook.save(path)
        workbook.close()
        return path

    def test_reordered_and_inserted_columns_are_found_by_header_name(self) -> None:
        template = self.create_template(
            ["CUSTOM INFO", "DESCRIPTION", "PART NUMBER", "QUANTITY"]
        )
        layout = converter.read_template_layout(template)
        output = self.root / "output.xlsx"

        converter.write_records_to_template(
            template,
            output,
            layout,
            [
                {"PART NUMBER": "P-100", "DESCRIPTION": "PLATE", "QUANTITY": 2}
            ],
            update_existing=False,
        )

        workbook = openpyxl.load_workbook(output)
        sheet = workbook["Parts List"]
        self.assertEqual(sheet["A5"].value, "CUSTOM INFO")
        self.assertEqual(sheet["B6"].value, "PLATE")
        self.assertEqual(sheet["C6"].value, "P-100")
        self.assertEqual(sheet["D6"].value, 2)
        workbook.close()

    def test_table_banding_extends_past_the_template_row_limit(self) -> None:
        template = self.create_template(
            ["PART NUMBER", "DESCRIPTION", "QUANTITY"],
            table_end_row=187,
        )
        layout = converter.read_template_layout(template)
        output = self.root / "output.xlsx"
        records = [
            {"PART NUMBER": f"P-{index:03d}", "DESCRIPTION": "PLATE"}
            for index in range(1, 191)
        ]

        converter.write_records_to_template(
            template,
            output,
            layout,
            records,
            update_existing=False,
        )

        workbook = openpyxl.load_workbook(output)
        sheet = workbook["Parts List"]
        table = sheet.tables["PartsListTable"]
        self.assertEqual(table.ref, "A5:C195")
        self.assertEqual(table.autoFilter.ref, "A5:C195")
        self.assertTrue(table.tableStyleInfo.showRowStripes)
        self.assertEqual(sheet["A195"].value, "P-190")
        workbook.close()

    def test_duplicate_normalized_template_headers_are_rejected(self) -> None:
        template = self.create_template(
            ["PART NUMBER", "PART-NUMBER", "DESCRIPTION"]
        )

        with self.assertRaisesRegex(ValueError, "duplicate column names"):
            converter.read_template_layout(template)


if __name__ == "__main__":
    unittest.main()
