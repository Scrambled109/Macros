import pandas as pd
import openpyxl
from copy import copy
import os
import re

# --- STEEL WEIGHT TO THICKNESS DICTIONARY ---
STEEL_WEIGHTS = {
    7.65: "3/16", 10.2: "1/4", 12.75: "5/16", 15.3: "3/8",
    17.85: "7/16", 20.4: "1/2", 22.95: "9/16", 25.5: "5/8",
    28.05: "11/16", 30.6: "3/4", 33.15: "13/16", 35.7: "7/8",
    40.8: "1", 45.9: "1 1/8", 51.0: "1 1/4", 56.1: "1 3/8",
    61.2: "1 1/2", 66.3: "1 5/8", 71.4: "1 3/4", 76.5: "1 7/8",
    81.6: "2", 86.7: "2 1/8", 91.8: "2 1/4", 102.0: "2 1/2",
    112.2: "2 3/4", 122.4: "3", 132.6: "3 1/4", 142.8: "3 1/2",
    153.0: "3 3/4", 163.2: "4", 173.4: "4 1/4", 183.6: "4 1/2",
    204.0: "5", 224.4: "5 1/2", 244.8: "6", 265.2: "6 1/2",
    285.6: "7", 306.0: "7 1/2", 326.4: "8", 367.2: "9", 408.0: "10"
}

def populate_parts_list(bom_filepath, template_filepath, output_filepath):
    print("Loading BOM data...")
    bom_df = pd.read_excel(bom_filepath, sheet_name='Lofting')

    print("Filtering for parts starting with 'DV'...")
    bom_df = bom_df[bom_df['ENG MAT ID'].astype(str).str.startswith('DS')]

    print("Removing duplicate routing steps...")
    bom_df = bom_df.drop_duplicates(subset=['ENG MAT ID'], keep='last')
    
    # --- CUSTOM PARSING FUNCTION ---
    def extract_desc_and_thickness(raw_string):
        desc = str(raw_string).upper()
        name = desc 
        thickness = ""
        
        parts = [p.strip() for p in desc.split(',')]
        known_shapes = ["PLATE", "FLAT BAR", "BAR", "TEE", "ANGLE", "BEAM", "TUBE", "PIPE"]
        
        for part in parts:
            for shape in known_shapes:
                if shape in part:
                    name = shape 
                    break
            if name != desc: 
                break
                
        for part in parts:
            if any(char.isdigit() for char in part) and any(ind in part for ind in ['#', 'T', 'X', 'W', '.']):
                if not any(shape in part for shape in known_shapes): 
                    
                    if '#' in part:
                        match = re.search(r'([\d\.]+)\s*#', part)
                        if match:
                            weight_val = round(float(match.group(1)), 2)
                            if weight_val in STEEL_WEIGHTS:
                                thickness = STEEL_WEIGHTS[weight_val]
                                break 
                                
                    thickness = part 
                    break
                    
        return name, thickness
    # -----------------------------------

    print("Parsing and mapping descriptions...")
    
    processed_rows = []
    for _, row in bom_df.iterrows():
        desc_name, thickness_shape = extract_desc_and_thickness(row['Description'])
        part_num = str(row['ENG MAT ID'])
        assembly_val = part_num.split('#')[0] if '#' in part_num else part_num
        
        row_dict = {
            'ORDER #': "",
            'PART NUMBER': part_num,
            'DESCRIPTION': desc_name,
            'QUANTITY': row['QTY'],
            'TOTAL QUANTITY': "",
            'THICKNESS/ SHAPE': thickness_shape,
            'WIDTH': row['Width'],
            'LENGTH': row['Length'],
            'MATERIAL TYPE': row['MTL Type'],
            'ASSEMBLY': assembly_val
        }
        processed_rows.append(row_dict)
        
    mapped_data = pd.DataFrame(processed_rows)

    if os.path.exists(output_filepath):
        print("Found an existing completed parts list! Opening to append data...")
        wb = openpyxl.load_workbook(output_filepath, data_only=False)
    else:
        print("No existing list found. Loading the blank Parts List template...")
        wb = openpyxl.load_workbook(template_filepath, data_only=False)

    sheet = wb.active 

    # Find the first completely empty row to start appending data safely
    start_row = 6 
    while sheet.cell(row=start_row, column=2).value is not None:
        start_row += 1

    print(f"Writing new mapped data starting at row {start_row}...")
    
    for index, row_data in mapped_data.iterrows():
        current_row = start_row + index
        
        # Mapping values cleanly into your exact template column coordinates
        sheet.cell(row=current_row, column=1).value = row_data['ORDER #']
        sheet.cell(row=current_row, column=2).value = row_data['PART NUMBER']
        sheet.cell(row=current_row, column=3).value = row_data['DESCRIPTION']
        sheet.cell(row=current_row, column=4).value = row_data['QUANTITY']
        sheet.cell(row=current_row, column=5).value = row_data['TOTAL QUANTITY']
        sheet.cell(row=current_row, column=6).value = row_data['THICKNESS/ SHAPE']
        sheet.cell(row=current_row, column=7).value = row_data['WIDTH']
        sheet.cell(row=current_row, column=8).value = row_data['LENGTH']
        sheet.cell(row=current_row, column=9).value = row_data['MATERIAL TYPE']
        sheet.cell(row=current_row, column=12).value = row_data['ASSEMBLY']

        # Fix bold formatting generated by default templates across columns A to N
        for col in range(1, 15):
            cell = sheet.cell(row=current_row, column=col)
            if cell.font:
                new_font = copy(cell.font)
                new_font.bold = False
                cell.font = new_font

    wb.save(output_filepath)
    print(f"Success! Parts list saved/updated at: {output_filepath}")

if __name__ == "__main__":
    BOM_FILE = 'G689D-0-4327X List of Parts A-BOM.xlsm'                
    TEMPLATE_FILE = 'TEMP.xlsx' 
    OUTPUT_FILE = 'BID18_4327_MASTERS_PARTS_LIST.xlsx' 
    populate_parts_list(BOM_FILE, TEMPLATE_FILE, OUTPUT_FILE)
